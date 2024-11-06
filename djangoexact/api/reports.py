from dataclasses import dataclass

import xlsxwriter.format
import xlsxwriter.worksheet
import api.models as api_models
import api.calculators as calculators
from typing import Optional
import math_model.no_time_dependency_final.ghg_emissions_classes as math_utils
import numpy as np
import xlsxwriter
from enum import Enum
import logging as log
import os
from django.conf import settings
import openpyxl as pxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Color, PatternFill, Font, Border
from datetime import datetime
from io import BytesIO
from rest_framework.test import APIRequestFactory

log.basicConfig(level=log.DEBUG)

# TODO: Calculations of submodules are a mess. Total cumulative and yearly emissions don't add up at all, so something is happening there. Probably double-counting. Need to refactor the way they are calculated before making them available in the reports.


class NotReadyError(Exception):
    pass


class Colors(Enum):
    LIGHT_ORANGE_HEX = "fce4d6"
    LIGHT_BLUE_HEX = "d9e1f2"
    LIGHT_BEIGE_HEX = "ddd9c4"
    LIGHT_ORANGE_FILL = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    LIGHT_BEIGE_FILL = PatternFill(start_color="ddd9c4", end_color="ddd9c4", fill_type="solid")


class ReportFactory:

    @staticmethod
    def get_report_class(module: api_models.Module | api_models.LandModule):
        if isinstance(module, api_models.PerennialCropland):
            return PerennialCroplandReport
        elif isinstance(module, api_models.AnnualCropland):
            return AnnualCroplandReport
        elif isinstance(module, api_models.FloodedRice):
            return FloodedRiceReport
        elif isinstance(module, api_models.LandUseChange):
            return LandUseChangeReport
        elif isinstance(module, api_models.SetAside):
            return SetAsideReport
        elif isinstance(module, api_models.Grassland):
            return GrasslandReport
        elif isinstance(module, api_models.OtherLand):
            return OtherLandReport
        elif isinstance(module, api_models.Waterbody):
            return WaterbodyReport
        elif isinstance(module, api_models.Aquaculture):
            return AquacultureReport
        elif isinstance(module, api_models.SmallFishery):
            return SmallFisheryReport
        elif isinstance(module, api_models.LargeFishery):
            return LargeFisheryReport
        elif isinstance(module, api_models.Livestock):
            return LivestockReport
        elif isinstance(module, api_models.ForestManagement):
            return ForestManagementReport
        elif isinstance(module, api_models.Energy):
            return EnergyReport
        elif isinstance(module, api_models.Input):
            return InputReport
        elif isinstance(module, api_models.Irrigation):
            return IrrigationReport
        elif isinstance(module, api_models.Settlement):
            return SettlementReport
        else:
            log.warning(f"No report class found for module {module.module_type.name}")
            return


class ExcelFileManager:
    def __init__(self):
        # Start with an empty in-memory Excel file
        self.excel_file = BytesIO()
        self._create_initial_excel()

    def _create_initial_excel(self):

        workbook = xlsxwriter.Workbook(self.excel_file, {"in_memory": True})
        workbook.add_worksheet("Results")
        workbook.add_worksheet("Metadata")
        workbook.add_worksheet("Additional Indicators")

        # Close the xlsxwriter workbook to finalize the file
        workbook.close()

        # Rewind the in-memory file
        self.excel_file.seek(0)

    def get_workbook(self):
        # Return an openpyxl workbook from the current in-memory file
        return pxl.load_workbook(self.excel_file)

    def save_workbook(self, workbook):
        # Save an openpyxl workbook back into the BytesIO stream
        self.excel_file = BytesIO()
        workbook.save(self.excel_file)
        self.excel_file.seek(0)

    def get_excel_bytes(self):
        # Get the current Excel file as bytes (e.g., for download)
        self.excel_file.seek(0)
        return self.excel_file.getvalue()


@dataclass
class BaseProjectReport:
    project: api_models.Project
    start_year_of_activities: int = None
    last_year_of_accounting: int = None
    implementation_years: int = None
    capitalization_years: int = None
    duration: int = None

    workbook: pxl.Workbook = None
    results_worksheet: Worksheet = None
    metadata_worksheet: Worksheet = None
    additional_indicators_worksheet: Worksheet = None
    activities: list[api_models.Activity] = None
    activity_reports: list["BaseActivityReport"] = None

    colors: Colors = Colors
    excel_manager: ExcelFileManager = None

    def __post_init__(self):
        self.activity_reports = []
        self.activities = self.project.activities.all()
        self.excel_manager = ExcelFileManager()
        pass

    def build_report_skeleton(self):
        """
        Builds the skeleton of a report for the project.

        This method initializes the report by setting up the necessary worksheets and populating them with
        headers and initial values. It retrieves the workbook from the excel manager, sets up the "Results",
        "Metadata", and "Additional Indicators" worksheets, and fills in the initial data for each worksheet.

        The following steps are performed:
        1. Logs the start of the report skeleton building process.
        2. Initializes the report's time-related attributes from the project.
        3. Retrieves the workbook and relevant worksheets from the excel manager.
        4. Sets up the headers and initial values in the "Results" worksheet.
        5. Sets up the headers and initial values in the "Metadata" worksheet.
        6. Sets up the headers and initial values in the "Additional Indicators" worksheet.
        7. Saves the workbook using the excel manager.
        8. Logs the completion of the report skeleton building process.

        Returns:
            Workbook: The workbook object with the initialized report skeleton.
        """
        log.debug(f"Building report skeleton for {self.project.name}")
        self.start_year_of_activities = self.project.start_year_of_activities
        self.last_year_of_accounting = self.project.last_year_of_accounting
        self.implementation_years = self.project.implementation_years
        self.capitalization_years = self.project.capitalization_years
        self.duration = self.implementation_years + self.capitalization_years

        self.workbook = self.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]
        self.metadata_worksheet = self.workbook["Metadata"]
        self.additional_indicators_worksheet = self.workbook["Additional Indicators"]

        self.results_worksheet.cell(row=1, column=1, value="Activity and GHGs / Years")

        self.results_worksheet.cell(row=2, column=1, value="Cumulative balance in Tco2-eq")
        self.results_worksheet.cell(row=3, column=1, value="Yearly balance in Tco2-eq")
        self.results_worksheet.cell(row=4, column=1, value="CO2 in biomass")
        self.results_worksheet.cell(row=5, column=1, value="CO2 in soils")
        self.results_worksheet.cell(row=6, column=1, value="Other CO2")
        self.results_worksheet.cell(row=7, column=1, value="CH4")
        self.results_worksheet.cell(row=8, column=1, value="N20")
        self.results_worksheet.cell(row=9, column=1, value="Other GHGs")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=1, column=i + 2, value=year)

        self.metadata_worksheet.cell(row=1, column=2, value="START")
        self.metadata_worksheet.cell(row=1, column=3, value="WITH")
        self.metadata_worksheet.cell(row=1, column=4, value="WITHOUT")

        self.additional_indicators_worksheet.cell(row=1, column=1, value="Activity and GHGs / Years")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.additional_indicators_worksheet.cell(row=1, column=i + 2, value=year)

        self.excel_manager.save_workbook(self.workbook)

        log.debug(f"Report skeleton for {self.project.name} built.")
        return self.workbook

    def finalize_report(self):
        """
        Finalizes the report for the project by aggregating emissions data from various activities and modules,
        calculating yearly and cumulative balances of CO2 equivalents, and writing the results to an Excel workbook.

        The method performs the following steps:
        1. Initializes lists to store emissions data for different greenhouse gases (GHGs) and CO2 sources.
        2. Iterates through activity reports and their respective module reports to extract emissions data.
        3. Aggregates the extracted emissions data for each type of GHG and CO2 source.
        4. Calculates the yearly and cumulative balance of CO2 equivalents.
        5. Writes the calculated balances and emissions data to the "Results" worksheet of an Excel workbook.
        6. Saves the updated workbook.

        The emissions data includes:
        - Other GHGs
        - N2O (Nitrous Oxide)
        - CH4 (Methane)
        - Other CO2 (excluding biomass and soil CO2 changes)
        - Soil CO2
        - Biomass CO2

        The results written to the Excel worksheet include:
        - Cumulative balance of CO2 equivalents for each year
        - Yearly balance of CO2 equivalents for each year
        - Emissions data for biomass CO2, soil CO2, other CO2, CH4, N2O, and other GHGs for each year

        Note: The method assumes that the `self.duration`, `self.activity_reports`, `self.project.name`,
        `self.start_year_of_activities`, `self.last_year_of_accounting`, `self.excel_manager`, and
        `self.workbook` attributes are properly initialized.
        """
        log.debug(f"Finalizing report for project {self.project.name}")

        other_ghgs = [[0] * self.duration]
        n2o = [[0] * self.duration]
        ch4 = [[0] * self.duration]
        other_co2 = [[0] * self.duration]
        soil_co2 = [[0] * self.duration]
        biomass_co2 = [[0] * self.duration]
        yearly_balance_t_co2_eq = [[0] * self.duration]
        cumulative_balance_t_co2_eq = [[0] * self.duration]

        for activity in self.activity_reports:
            log.debug(f"Finalizing report for activity {activity.activity_title}")
            for module in activity.modules_reports:
                log.debug(f"Finalizing report for module {module.module_title}")
                other_ghgs.append(module.extract_emissions(module.emissions_set, gas_type=math_utils.GasTypes.OTHER))
                n2o.append(module.extract_emissions(module.emissions_set, gas_type=math_utils.GasTypes.N2O))
                ch4.append(module.extract_emissions(module.emissions_set, gas_type=math_utils.GasTypes.CH4))
                other_co2.append(module.extract_emissions(module.emissions_set, activity_type=None, gas_type=math_utils.GasTypes.CO2, excluded_activity_types=[math_utils.ActivityTypes.BIOMASS, math_utils.ActivityTypes.SOIL_CO2_CHANGE]))
                soil_co2.append(module.extract_emissions(module.emissions_set, activity_type=math_utils.ActivityTypes.SOIL_CO2_CHANGE, gas_type=math_utils.GasTypes.CO2))
                biomass_co2.append(module.extract_emissions(module.emissions_set, activity_type=math_utils.ActivityTypes.BIOMASS, gas_type=math_utils.GasTypes.CO2))

        other_ghgs = list(map(sum, zip(*other_ghgs)))
        n2o = list(map(sum, zip(*n2o)))
        ch4 = list(map(sum, zip(*ch4)))
        other_co2 = list(map(sum, zip(*other_co2)))
        soil_co2 = list(map(sum, zip(*soil_co2)))
        biomass_co2 = list(map(sum, zip(*biomass_co2)))

        yearly_balance_t_co2_eq = list(map(sum, zip(biomass_co2, soil_co2, other_co2, ch4, n2o, other_ghgs)))
        cumulative_balance_t_co2_eq = np.cumsum(yearly_balance_t_co2_eq)

        self.workbook = self.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=2, column=i + 2, value=cumulative_balance_t_co2_eq[i])
            self.results_worksheet.cell(row=3, column=i + 2, value=yearly_balance_t_co2_eq[i])
            self.results_worksheet.cell(row=4, column=i + 2, value=biomass_co2[i])
            self.results_worksheet.cell(row=5, column=i + 2, value=soil_co2[i])
            self.results_worksheet.cell(row=6, column=i + 2, value=other_co2[i])
            self.results_worksheet.cell(row=7, column=i + 2, value=ch4[i])
            self.results_worksheet.cell(row=8, column=i + 2, value=n2o[i])
            self.results_worksheet.cell(row=9, column=i + 2, value=other_ghgs[i])

        self.excel_manager.save_workbook(self.workbook)

    def build_report(self):
        """
        Builds the complete report by following these steps:

        1. Initializes the report structure by calling `build_report_skeleton`.
        2. Iterates over each activity in `self.activities`:
            - Creates an instance of `BaseActivityReport` for the activity.
            - Builds the skeleton for the activity report.
            - Builds the module reports for the activity.
            - Appends the activity report to `self.activity_reports`.
        3. Finalizes the report by calling `finalize_report`.

        Returns:
            tuple: A tuple containing the Excel file memory pointer and its byte representation.
        """
        self.build_report_skeleton()

        for activity in self.activities:
            activity_report = BaseActivityReport(self, activity)
            activity_report.build_activity_skeleton()
            activity_report.build_modules_reports()
            self.activity_reports.append(activity_report)

        self.finalize_report()
        return self.excel_manager.excel_file, self.excel_manager.get_excel_bytes()

    def close_file(self):
        """
        Closes the Excel file managed by the excel_manager.

        This method ensures that the Excel file associated with the excel_manager
        is properly closed, releasing any in-memory resources held by the file.
        """
        self.excel_manager.excel_file.close()


@dataclass
class BaseActivityReport:

    project_report: BaseProjectReport
    activity: api_models.Activity
    modules_reports: list["BaseModuleReport"] = None

    activity_title: str = None
    start_year_of_activities: int = None
    last_year_of_accounting: int = None
    implementation_years: int = None
    capitalization_years: int = None
    duration: int = None

    workbook: pxl.Workbook = None
    results_worksheet: Worksheet = None
    metadata_worksheet: Worksheet = None
    additional_indicators_worksheet: Worksheet = None

    def __post_init__(self):
        self.modules_reports = []

    def build_activity_skeleton(self):
        """
        Builds the activity skeleton for the current activity and updates the Excel workbook with relevant data.

        This method performs the following steps:
        1. Logs the start of the activity skeleton building process.
        2. Sets various attributes related to the activity, such as title, start year, last year of accounting,
           implementation years, capitalization years, and total duration.
        3. Retrieves the workbook and specific worksheets (Results, Metadata, Additional Indicators) from the
           project report's Excel manager.
        4. Determines the last row in each worksheet to append new data.
        5. Logs the last row numbers for each worksheet.
        6. Updates the Results, Metadata, and Additional Indicators worksheets with the activity title and
           specific values.
        7. Saves the updated workbook using the project report's Excel manager.
        8. Logs the completion of the activity skeleton building process.

        Returns:
            Workbook: The updated Excel workbook.
        """
        log.debug(f"Building activity skeleton for {self.activity.name}")
        self.activity_title = self.activity.name
        self.start_year_of_activities = self.activity.project.start_year_of_activities
        self.last_year_of_accounting = self.activity.project.last_year_of_accounting
        self.implementation_years = self.activity.implementation_years
        self.capitalization_years = self.activity.capitalization_years
        self.duration = self.implementation_years + self.capitalization_years

        self.workbook = self.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]
        self.metadata_worksheet = self.workbook["Metadata"]
        self.additional_indicators_worksheet = self.workbook["Additional Indicators"]

        last_results_row = self.results_worksheet.max_row + 1
        last_metadata_row = self.metadata_worksheet.max_row + 1
        last_additional_indicators_row = self.additional_indicators_worksheet.max_row + 1
        log.debug(f"Last results row: {last_results_row}")
        log.debug(f"Last metadata row: {last_metadata_row}")
        log.debug(f"Last additional indicators row: {last_additional_indicators_row}")

        self.results_worksheet.cell(row=last_results_row, column=1, value=str(self.activity_title)[:6])
        self.results_worksheet.cell(row=last_results_row, column=1).fill = Colors.LIGHT_ORANGE_FILL.value

        self.metadata_worksheet.cell(row=last_metadata_row, column=1, value=str(self.activity_title)[:6])
        self.metadata_worksheet.cell(row=last_metadata_row, column=1).fill = Colors.LIGHT_BLUE_FILL.value

        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row, column=1, value=str(self.activity_title)[:6])
        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row, column=1).fill = Colors.LIGHT_ORANGE_FILL.value

        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row + 1, column=1, value="Land Uses Targeted (ha)")
        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row + 1, column=1).fill = Colors.LIGHT_BLUE_FILL.value

        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row + 2, column=1, value="With project")
        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row + 2, column=1, value="With project").fill = Colors.LIGHT_BEIGE_FILL.value
        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row + 3, column=1, value="Without project")
        self.additional_indicators_worksheet.cell(row=last_additional_indicators_row + 3, column=1, value="Without project").fill = Colors.LIGHT_BEIGE_FILL.value

        self.project_report.excel_manager.save_workbook(self.workbook)

        log.debug(f"Activity skeleton for {self.activity.name} built.")

        return self.workbook

    def build_modules_reports(self) -> list["BaseModuleReport"]:
        """
        Builds reports for each module in the activity.

        This method iterates over the modules associated with the activity,
        retrieves the appropriate report class for each module using the
        ReportFactory, and builds the report. If no report class is found
        for a module, a warning is logged. The generated reports are appended
        to the `modules_reports` list.

        Returns:
            list[BaseModuleReport]: A list of generated module reports.
        """
        modules = self.activity.modules
        for module in modules:
            log.debug(f"Building report for module {module.module_type.name}")
            ReportClass = ReportFactory.get_report_class(module)
            if ReportClass is None:
                log.warning(f"No report class found for module {module.module_type.name}")
                continue
            module_report = ReportClass(module, self)
            module_report.build_report()
            self.modules_reports.append(module_report)

        return self.modules_reports

    def extract_modules_emissions(self, activity_type=None, gas_type=None) -> list[float]:
        """
        Extracts and aggregates emissions from module reports based on the specified activity type and gas type.

        Args:
            activity_type (str, optional): The type of activity to filter emissions by. Defaults to None.
            gas_type (str, optional): The type of gas to filter emissions by. Defaults to None.

        Returns:
            list[float]: A list of aggregated emissions values for the specified duration.
        """
        emissions = np.zeros(self.duration)
        for module_report in self.modules_reports:
            emissions += module_report.extract_emissions(module_report.emissions_set, activity_type, gas_type)
        return emissions


@dataclass
class BaseModuleReport:
    module: api_models.Module | api_models.LandModule
    activity_report: BaseActivityReport = None
    module_title: str = None
    units: float = None
    calculator: calculators.BaseCalculator | calculators.LandModuleCalculator = None
    result: dict = None
    emissions_set: list[math_utils.YearlyGasActivityEmissionSet] = None

    start_year_of_activities: int = None
    last_year_of_accounting: int = None
    implementation_years: int = None
    capitalization_years: int = None
    duration: int = None

    workbook: pxl.Workbook = None
    results_worksheet: Worksheet = None
    metadata_worksheet: Worksheet = None
    additional_indicators_worksheet: Worksheet = None

    def __post_init__(self):

        try:
            self.result = self.calculator.calculate()
        except Exception as e:
            log.error(f"Cannot calculate report for module {self.module.module_type.name} in activity {self.module.activity.name}: {e}")
            raise NotReadyError(f"Cannot calculate report for module {self.module.module_type.name} in activity {self.module.activity.name}: {e}")

        from api.calculators import Result

        self.emissions_set = Result(*self.result).balance.yearly_emissions_by_sector_by_gas

        self.start_year_of_activities = self.module.activity.project.start_year_of_activities
        self.last_year_of_accounting = self.module.activity.project.last_year_of_accounting
        self.implementation_years = self.module.activity.implementation_years
        self.capitalization_years = self.module.activity.capitalization_years
        self.duration = self.module.activity.implementation_years + self.module.activity.capitalization_years

    def get_result(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]
        self.metadata_worksheet = self.workbook["Metadata"]
        self.additional_indicators_worksheet = self.workbook["Additional Indicators"]
        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)

    def build_report(self):
        """
        Builds and updates the report by adding module type information to the results worksheet.

        This method performs the following steps:
        1. Retrieves the workbook from the excel manager.
        2. Accesses the "Results" worksheet within the workbook.
        3. Determines the next available row in the "Results" worksheet.
        4. Inserts the module type name into the first column of the next available row.
        5. Applies a light blue fill to the cell containing the module type name.
        6. Saves the updated workbook using the excel manager.

        Attributes:
            workbook (Workbook): The Excel workbook object.
            results_worksheet (Worksheet): The worksheet where results are recorded.
            last_results_row (int): The next available row in the results worksheet.
        """
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row + 1
        self.results_worksheet.cell(row=last_results_row, column=1, value=str(self.module.module_type.name))
        self.results_worksheet.cell(row=last_results_row, column=1).fill = Colors.LIGHT_BLUE_FILL.value

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)

    def extract_emissions(self, data, activity_type=None, gas_type=None, excluded_activity_types=[], excluded_gas_types=[]) -> np.ndarray:
        """
        Extracts emissions values from the provided data based on the specified activity type and gas type.

        Args:
            data (list): A list of data entries, where each entry is expected to have 'activity', 'gas_type', and 'emissions' attributes.
            activity_type (str, optional): The type of activity to filter the data entries.
            gas_type (str, optional): The type of gas to filter the data entries.

        Returns:
            list: A list of emission values if a matching entry is found.
            numpy.ndarray: An array of zeros with a length equal to the sum of implementation years and capitalization years if no matching entry is found.
        """
        emissions = [[0] * self.duration]

        if activity_type is not None and gas_type is not None:
            for entry in data:
                if entry.activity == activity_type and entry.activity not in excluded_activity_types and entry.gas_type == gas_type and entry.gas_type not in excluded_gas_types:
                    log.debug(f"Found emissions for {activity_type} and {gas_type} excluding {excluded_activity_types} and {excluded_gas_types}")
                    entry_emissions = [e.value for e in entry.emissions]
                    emissions.append(entry_emissions)

        elif activity_type is not None:
            for entry in data:
                if entry.activity == activity_type and entry.activity not in excluded_activity_types and entry.gas_type not in excluded_gas_types:
                    log.debug(f"Found emissions for {entry.activity}, {entry.gas_type} excluding {excluded_gas_types}")
                    entry_emissions = [e.value for e in entry.emissions]
                    emissions.append(entry_emissions)

        elif gas_type is not None and gas_type not in excluded_gas_types:
            for entry in data:
                if entry.gas_type == gas_type and entry.gas_type not in excluded_gas_types and entry.activity not in excluded_activity_types:
                    log.debug(f"Found emissions for {entry.activity}, {entry.gas_type} excluding {excluded_activity_types}")
                    entry_emissions = [e.value for e in entry.emissions]
                    emissions.append(entry_emissions)

        else:
            log.debug("No activity or gas type specified. Extracting all emissions.")
            for entry in data:
                entry_emissions = [e.value for e in entry.emissions]
                emissions.append(entry_emissions)

        summed_emissions = list(map(sum, zip(*emissions)))

        return summed_emissions


@dataclass
class LandModuleReport(BaseModuleReport):

    biomass_co2: list[float] = None
    soil_co2: list[float] = None
    soil_n2o: list[float] = None
    fire_n2o: list[float] = None
    fire_ch4: list[float] = None
    emissions_set: list[math_utils.YearlyGasActivityEmissionSet] = None

    biomass_co2_row_index: int = None
    soil_co2_row_index: int = None
    soil_n2o_row_index: int = None
    fire_n2o_row_index: int = None
    fire_ch4_row_index: int = None

    biomass_co2_source = (math_utils.ActivityTypes.BIOMASS, math_utils.GasTypes.CO2)
    soil_co2_source = (math_utils.ActivityTypes.SOIL_CO2_CHANGE, math_utils.GasTypes.CO2)
    soil_n2o_source = (math_utils.ActivityTypes.SOM, math_utils.GasTypes.N2O)
    fire_n2o_source = (math_utils.ActivityTypes.RESIDUE_BURNING, math_utils.GasTypes.N2O)
    fire_ch4_source = (math_utils.ActivityTypes.RESIDUE_BURNING, math_utils.GasTypes.CH4)

    units_breakdown: list[float] = None

    units_breakdown_w: list[float] = None
    units_breakdown_wo: list[float] = None

    def __post_init__(self):
        """
        Post-initialization method to set up hectares breakdown calculations.

        This method is called after the object's initialization to compute the
        breakdown of hectares with and without certain conditions. It retrieves
        the necessary data from the calculator attributes and combines them to
        form the units breakdown.

        Attributes:
            hectares_length (int): The total length of the hectares array,
                calculated from the sum of implementation years and capitalization years.
            break_start_w (numpy.ndarray): Initial hectares data with conditions.
            break_start_wo (numpy.ndarray): Initial hectares data without conditions.
            break_w (numpy.ndarray): Hectares data with conditions.
            break_wo (numpy.ndarray): Hectares data without conditions.
            units_breakdown_w (list): Combined hectares data with conditions.
            units_breakdown_wo (list): Combined hectares data without conditions.
        """
        super().__post_init__()

        hectares_length = self.module.activity.implementation_years + self.module.activity.capitalization_years

        break_start_w = getattr(self.calculator.math_start_w, "hectares_total", np.zeros(hectares_length))
        break_start_wo = getattr(self.calculator.math_start_wo, "hectares_total", np.zeros(hectares_length))
        break_w = getattr(self.calculator.math_w, "hectares_total", np.zeros(hectares_length))
        break_wo = getattr(self.calculator.math_wo, "hectares_total", np.zeros(hectares_length))

        self.units_breakdown_w = [x + y for x, y in zip(break_start_w, break_w)]
        self.units_breakdown_wo = [x + y for x, y in zip(break_start_wo, break_wo)]

    def get_result(self):
        """
        Computes and assigns various emission values to the instance variables.

        This method extracts emission values for biomass CO2, soil CO2, soil N2O, fire N2O, and fire CH4
        from the emissions set using the specified sources. It also assigns the module title and units
        based on the module's type and area.

        Attributes:
            module_title (str): The title of the module based on its type.
            units (str): The area of the module.
            biomass_co2 (float): The extracted biomass CO2 emissions.
            soil_co2 (float): The extracted soil CO2 emissions.
            soil_n2o (float): The extracted soil N2O emissions.
            fire_n2o (float): The extracted fire N2O emissions.
            fire_ch4 (float): The extracted fire CH4 emissions.
        """
        self.module_title = self.module.module_type.name
        self.units = self.module.area

        self.biomass_co2 = self.extract_emissions(self.emissions_set, self.biomass_co2_source[0], self.biomass_co2_source[1])
        self.soil_co2 = self.extract_emissions(self.emissions_set, self.soil_co2_source[0], self.soil_co2_source[1])
        self.soil_n2o = self.extract_emissions(self.emissions_set, self.soil_n2o_source[0], self.soil_n2o_source[1])
        self.fire_n2o = self.extract_emissions(self.emissions_set, self.fire_n2o_source[0], self.fire_n2o_source[1])
        self.fire_ch4 = self.extract_emissions(self.emissions_set, self.fire_ch4_source[0], self.fire_ch4_source[1])

    def build_report(self):
        """
        Builds a detailed report by populating an Excel workbook with emissions data.

        This method performs the following steps:
        1. Calls the superclass's build_report method.
        2. Logs the start of the report building process.
        3. Retrieves the result data.
        4. Obtains the workbook and the "Results" worksheet.
        5. Determines the last row in the "Results" worksheet.
        6. Writes emissions information labels in the worksheet.
        7. Stores the row indices for different emissions data.
        8. Populates the worksheet with emissions data for each year in the specified range.
        9. Saves the updated workbook.
        10. Logs the completion of the report building process.

        Returns:
            openpyxl.Workbook: The populated Excel workbook.
        """
        super().build_report()
        log.debug(f"Building base report for {self.module.module_type.name}")
        self.get_result()

        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        # Write emissions information
        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CO2 in biomass")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CO2 in soils")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="N2O in soils")
        self.results_worksheet.cell(row=last_results_row + 4, column=1, value="N2O from fires")
        self.results_worksheet.cell(row=last_results_row + 5, column=1, value="CH4 from fires")

        self.biomass_co2_row_index = last_results_row + 1
        self.soil_co2_row_index = last_results_row + 2
        self.soil_n2o_row_index = last_results_row + 3
        self.fire_n2o_row_index = last_results_row + 4
        self.fire_ch4_row_index = last_results_row + 5

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.biomass_co2[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.soil_co2[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.soil_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 4, column=i + 2, value=self.fire_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 5, column=i + 2, value=self.fire_ch4[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)

        log.debug(f"Base report for {self.module.module_type.name} built.")
        return self.workbook


@dataclass
class LandUseChangeReport(LandModuleReport):
    module: api_models.LandUseChange

    dom_co2_source = (math_utils.ActivityTypes.DOM, math_utils.GasTypes.CO2)

    def __post_init__(self):
        self.calculator = calculators.LandUseChangeCalculator(self.module)
        return super().__post_init__()

    def get_result(self):
        super().get_result()

        dom_co2 = self.extract_emissions(self.emissions_set, self.dom_co2_source[0], self.dom_co2_source[1])
        self.soil_co2 = list(map(sum, zip(self.soil_co2, dom_co2)))


@dataclass
class PerennialCroplandReport(LandModuleReport):

    module: api_models.PerennialCropland

    def __post_init__(self):
        self.calculator = calculators.PerennialCroplandCalculator(self.module)
        return super().__post_init__()

    def populate_metadata(self):
        """
        Populates the metadata worksheet in the Excel workbook with relevant data.

        This method retrieves the workbook and the "Metadata" worksheet, then appends
        metadata information related to perennial cropland, hectares, main season crop,
        tillage management type, organic input type, and yield. The data is populated
        based on the module's state (start, with, without).

        The method performs the following steps:
        1. Retrieves the workbook and the "Metadata" worksheet.
        2. Determines the last row in the metadata worksheet.
        3. Adds metadata headers and fills the first column with static values.
        4. Populates the second, third, and fourth columns with dynamic values based on
           the module's state (start, with, without).
        5. Saves the updated workbook.

        The columns are populated as follows:
        - Column 1: Static metadata headers.
        - Column 2: Values when the module is in the start state.
        - Column 3: Values when the module is in the with state.
        - Column 4: Values when the module is in the without state.

        The method ensures that if certain values are not available, default values are used.

        Raises:
            AttributeError: If any required attributes are missing from the module or activity report.
        """
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.metadata_worksheet = self.workbook["Metadata"]

        last_metadata_row = self.metadata_worksheet.max_row + 1

        self.metadata_worksheet.cell(row=last_metadata_row, column=1, value="Perennial Cropland")
        self.metadata_worksheet.cell(row=last_metadata_row, column=1, value="Perennial Cropland").fill = Colors.LIGHT_BLUE_FILL.value

        self.metadata_worksheet.cell(row=last_metadata_row + 1, column=1, value="Hectares")
        self.metadata_worksheet.cell(row=last_metadata_row + 2, column=1, value="Main season crop")
        self.metadata_worksheet.cell(row=last_metadata_row + 3, column=1, value="Tillage management type")
        self.metadata_worksheet.cell(row=last_metadata_row + 4, column=1, value="Organic input type")
        self.metadata_worksheet.cell(row=last_metadata_row + 5, column=1, value="Yield")

        if self.module.is_start():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=2, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=2, value=self.module.land_use_type_start.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 3, column=2, value=self.module.tillage_management_type_start.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 4, column=2, value=self.module.organic_input_type_start.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 5, column=2, value=self.module.crop_yield_t2_start if self.module.crop_yield_t2_start is not None else "Default")

        if self.module.is_with():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=3, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=3, value=self.module.land_use_type_w.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 3, column=3, value=self.module.tillage_management_type_w.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 4, column=3, value=self.module.organic_input_type_w.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 5, column=3, value=self.module.crop_yield_t2_w if self.module.crop_yield_t2_w is not None else "Default")

        if self.module.is_without():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=4, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=4, value=self.module.land_use_type_wo.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 3, column=4, value=self.module.tillage_management_type_wo.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 4, column=4, value=self.module.organic_input_type_wo.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 5, column=4, value=self.module.crop_yield_t2_wo if self.module.crop_yield_t2_wo is not None else "Default")

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)

    def populate_additional_indicators(self):
        """
        Populates the 'Additional Indicators' worksheet in the Excel workbook with data for both
        'With project' and 'Without project' scenarios.

        This method performs the following steps:
        1. Retrieves the workbook and the 'Additional Indicators' worksheet.
        2. Finds the row corresponding to the 'With project' label and inserts a new row below it.
        3. Populates the new row with 'Perennial Cropland (ha)' and corresponding yearly data for the 'With project' scenario.
        4. Finds the row corresponding to the 'Without project' label and inserts a new row below it.
        5. Populates the new row with 'Perennial Cropland (ha)' and corresponding yearly data for the 'Without project' scenario.
        6. Saves the updated workbook.

        Attributes:
            workbook (Workbook): The Excel workbook object.
            additional_indicators_worksheet (Worksheet): The worksheet for additional indicators.
            with_project_row (int): The row number for the 'With project' label.
            without_project_row (int): The row number for the 'Without project' label.
            start_year_of_activities (int): The starting year of activities.
            last_year_of_accounting (int): The last year of accounting.
            units_breakdown_w (list): The list of units breakdown for 'With project' scenario.
            units_breakdown_wo (list): The list of units breakdown for 'Without project' scenario.
        """
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.additional_indicators_worksheet = self.workbook["Additional Indicators"]

        with_project_row = None
        for row in self.additional_indicators_worksheet.iter_rows():
            if row[0].value == "With project":
                with_project_row = row[0].row

        self.additional_indicators_worksheet.insert_rows(with_project_row + 1)
        with_project_row = with_project_row + 1

        self.additional_indicators_worksheet.cell(with_project_row, 1, "Perennial Cropland (ha)")
        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.additional_indicators_worksheet.cell(with_project_row, i + 2, self.units_breakdown_w[i])

        without_project_row = None
        for row in self.additional_indicators_worksheet.iter_rows():
            if row[0].value == "Without project":
                without_project_row = row[0].row

        self.additional_indicators_worksheet.insert_rows(without_project_row + 1)
        without_project_row = without_project_row + 1

        self.additional_indicators_worksheet.cell(without_project_row, 1, "Perennial Cropland (ha)")
        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.additional_indicators_worksheet.cell(without_project_row, i + 2, self.units_breakdown_wo[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)

    def add_minor_seasons_results(self):
        """
        Adds the results of minor seasons to the current module.

        This method iterates through all minor seasons associated with the module and calculates
        the emissions for each minor season using the PerennialCropCalculator. The emissions are
        then aggregated into the current module's biomass CO2, soil CO2, soil N2O, fire N2O, and
        fire CH4 attributes.

        The method handles both "with" and "without" scenarios based on the module's configuration.

        Attributes:
            minor_seasons (list): List of minor seasons associated with the module.
            minor_calculator (PerennialCropCalculator): Calculator instance for the minor season.
            minor_emission_set (list): List to store the emissions for the minor season.

        Emissions are extracted and added to the following attributes:
            self.biomass_co2 (list): Aggregated biomass CO2 emissions.
            self.soil_co2 (list): Aggregated soil CO2 emissions.
            self.soil_n2o (list): Aggregated soil N2O emissions.
            self.fire_n2o (list): Aggregated fire N2O emissions.
            self.fire_ch4 (list): Aggregated fire CH4 emissions.
        """
        minor_seasons = getattr(self.module, "minor_seasons", [])

        for minor_season in minor_seasons.all():
            log.debug(f"Building report for minor season {minor_season.name}")

            minor_calculator = calculators.PerennialCropCalculator(minor_season)
            minor_emission_set = []

            if self.module.is_with():
                minor_emission_set += minor_calculator.results_w.yearly_emissions_by_sector_by_gas
                if self.calculator.results_start_w is not None:
                    minor_emission_set += self.calculator.results_start_w.yearly_emissions_by_sector_by_gas

            if self.module.is_without():
                minor_emission_set += minor_calculator.results_wo.yearly_emissions_by_sector_by_gas
                if self.calculator.results_start_wo is not None:
                    minor_emission_set += minor_calculator.results_start_wo.yearly_emissions_by_sector_by_gas

            self.biomass_co2 = [x + y for x, y in zip(self.biomass_co2, self.extract_emissions(minor_emission_set, self.biomass_co2_source[0], self.biomass_co2_source[1]))]
            self.soil_co2 = [x + y for x, y in zip(self.soil_co2, self.extract_emissions(minor_emission_set, self.soil_co2_source[0], self.soil_co2_source[1]))]
            self.soil_n2o = [x + y for x, y in zip(self.soil_n2o, self.extract_emissions(minor_emission_set, self.soil_n2o_source[0], self.soil_n2o_source[1]))]
            self.fire_n2o = [x + y for x, y in zip(self.fire_n2o, self.extract_emissions(minor_emission_set, self.fire_n2o_source[0], self.fire_n2o_source[1]))]
            self.fire_ch4 = [x + y for x, y in zip(self.fire_ch4, self.extract_emissions(minor_emission_set, self.fire_ch4_source[0], self.fire_ch4_source[1]))]

    def get_result(self):
        """
        Generates the result for the report by building and populating it with necessary data.

        This method performs the following steps:
        1. Calls the superclass's `get_result` method.
        2. Logs the start of the report building process.
        3. Adds minor seasons results to the report.
        4. Populates the report with metadata.
        5. Populates the report with additional indicators.
        6. Logs the completion of the report building process.
        7. Returns the generated report as Excel bytes.

        Returns:
            bytes: The generated report in Excel format.
        """
        super().get_result()
        log.debug(f"Building report for {self.module.module_type.name}")

        self.add_minor_seasons_results()
        self.populate_metadata()
        self.populate_additional_indicators()

        log.debug(f"Report for {self.module_title} built.")
        return self.activity_report.project_report.excel_manager.get_excel_bytes()


@dataclass
class AnnualCroplandReport(LandModuleReport):

    module: api_models.AnnualCropland
    activity_report: BaseActivityReport = None

    def __post_init__(self):
        self.calculator = calculators.AnnualCroplandCalculator(self.module)
        return super().__post_init__()

    def add_minor_seasons_results(self):
        minor_seasons = getattr(self.module, "submodules", [])

        for minor_season in minor_seasons:
            log.debug(f"Building report for minor season {minor_season.name}")

            minor_calculator = calculators.AnnualCropCalculator(minor_season)
            minor_emission_set = []

            if self.module.is_with():
                minor_emission_set += minor_calculator.results_w.yearly_emissions_by_sector_by_gas
                if self.calculator.results_start_w is not None:
                    minor_emission_set += self.calculator.results_start_w.yearly_emissions_by_sector_by_gas

            if self.module.is_without():
                minor_emission_set += minor_calculator.results_wo.yearly_emissions_by_sector_by_gas
                if self.calculator.results_start_wo is not None:
                    minor_emission_set += self.calculator.results_start_wo.yearly_emissions_by_sector_by_gas

            self.biomass_co2 = [x + y for x, y in zip(self.biomass_co2, self.extract_emissions(minor_emission_set, self.biomass_co2_source[0], self.biomass_co2_source[1]))]
            self.soil_co2 = [x + y for x, y in zip(self.soil_co2, self.extract_emissions(minor_emission_set, self.soil_co2_source[0], self.soil_co2_source[1]))]
            self.soil_n2o = [x + y for x, y in zip(self.soil_n2o, self.extract_emissions(minor_emission_set, self.soil_n2o_source[0], self.soil_n2o_source[1]))]
            self.fire_n2o = [x + y for x, y in zip(self.fire_n2o, self.extract_emissions(minor_emission_set, self.fire_n2o_source[0], self.fire_n2o_source[1]))]
            self.fire_ch4 = [x + y for x, y in zip(self.fire_ch4, self.extract_emissions(minor_emission_set, self.fire_ch4_source[0], self.fire_ch4_source[1]))]

    def build_metadata(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.metadata_worksheet = self.workbook["Metadata"]

        last_metadata_row = self.metadata_worksheet.max_row + 1

        self.metadata_worksheet.cell(row=last_metadata_row, column=1, value="Annual Cropland")
        self.metadata_worksheet.cell(row=last_metadata_row, column=1, value="Annual Cropland").fill = Colors.LIGHT_BLUE_FILL.value

        self.metadata_worksheet.cell(row=last_metadata_row + 1, column=1, value="Hectares")
        self.metadata_worksheet.cell(row=last_metadata_row + 2, column=1, value="Main season crop")
        self.metadata_worksheet.cell(row=last_metadata_row + 3, column=1, value="Tillage management type")
        self.metadata_worksheet.cell(row=last_metadata_row + 4, column=1, value="Organic input type")
        self.metadata_worksheet.cell(row=last_metadata_row + 5, column=1, value="Residue management type")
        self.metadata_worksheet.cell(row=last_metadata_row + 6, column=1, value="Yield")
        self.metadata_worksheet.cell(row=last_metadata_row + 7, column=1, value="Minor season crop")
        self.metadata_worksheet.cell(row=last_metadata_row + 8, column=1, value="Minor residue management type")
        self.metadata_worksheet.cell(row=last_metadata_row + 9, column=1, value="Minor yield")

        if self.module.is_start():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=2, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=2, value=self.module.land_use_type_start.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 3, column=2, value=self.module.tillage_management_type_start.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 4, column=2, value=self.module.organic_input_type_start.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 5, column=2, value=self.module.residue_management_type_start.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 6, column=2, value=self.module.crop_yield_t2_start if self.module.crop_yield_t2_start is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 7, column=2, value=self.module.minor_land_use_type_start.name if self.module.minor_land_use_type_start is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 8, column=2, value=self.module.minor_residue_management_type_start.name if self.module.minor_residue_management_type_start is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 9, column=2, value=self.module.minor_yield_start if self.module.minor_yield_start is not None else "Default")

        if self.module.is_with():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=3, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=3, value=self.module.land_use_type_w.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 3, column=3, value=self.module.tillage_management_type_w.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 4, column=3, value=self.module.organic_input_type_w.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 5, column=3, value=self.module.residue_management_type_w.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 6, column=3, value=self.module.crop_yield_t2_w if self.module.crop_yield_t2_w is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 7, column=3, value=self.module.minor_land_use_type_w.name if self.module.minor_land_use_type_w is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 8, column=3, value=self.module.minor_residue_management_type_w.name if self.module.minor_residue_management_type_w is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 9, column=3, value=self.module.minor_yield_w if self.module.minor_yield_w is not None else "Default")

        if self.module.is_without():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=4, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=4, value=self.module.land_use_type_wo.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 3, column=4, value=self.module.tillage_management_type_wo.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 4, column=4, value=self.module.organic_input_type_wo.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 5, column=4, value=self.module.residue_management_type_wo.name)
            self.metadata_worksheet.cell(row=last_metadata_row + 6, column=4, value=self.module.crop_yield_t2_wo if self.module.crop_yield_t2_wo is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 7, column=4, value=self.module.minor_land_use_type_wo.name if self.module.minor_land_use_type_wo is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 8, column=4, value=self.module.minor_residue_management_type_wo.name if self.module.minor_residue_management_type_wo is not None else "Default")
            self.metadata_worksheet.cell(row=last_metadata_row + 9, column=4, value=self.module.minor_yield_wo if self.module.minor_yield_wo is not None else "Default")

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)

    def build_additional_indicators(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.additional_indicators_worksheet = self.workbook["Additional Indicators"]

        with_project_row = None
        for row in self.additional_indicators_worksheet.iter_rows():
            if row[0].value == "With project":
                with_project_row = row[0].row

        self.additional_indicators_worksheet.insert_rows(with_project_row + 1)
        with_project_row = with_project_row + 1

        self.additional_indicators_worksheet.cell(with_project_row, 1, "Annual Cropland (ha)")
        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.additional_indicators_worksheet.cell(with_project_row, i + 2, self.units_breakdown_w[i])

        without_project_row = None
        for row in self.additional_indicators_worksheet.iter_rows():
            if row[0].value == "Without project":
                without_project_row = row[0].row

        self.additional_indicators_worksheet.insert_rows(without_project_row + 1)
        without_project_row = without_project_row + 1

        self.additional_indicators_worksheet.cell(without_project_row, 1, "Annual Cropland (ha)")
        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.additional_indicators_worksheet.cell(without_project_row, i + 2, self.units_breakdown_wo[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)

    def get_result(self):
        """
        # TODO: Additional indicators sheet: hectars indicator is self.hectares_total. Livestock indicator under self.livestock_heads_yearly_breakdown. Catch (fish) under self.tonnes_catch_yearly_breakdown
        """
        super().get_result()

        log.debug(f"Building report for {self.module.module_type.name}")

        self.add_minor_seasons_results()
        self.build_metadata()
        self.build_additional_indicators()

        log.debug(f"Report for {self.module_title} built.")

        return self.workbook


@dataclass
class SetAsideReport(LandModuleReport):

    module: api_models.SetAside
    activity_report: BaseActivityReport = None

    def __post_init__(self):
        self.calculator = calculators.SetAsideCalculator(self.module)
        return super().__post_init__()

    def get_result(self):
        super().get_result()

        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]
        self.metadata_worksheet = self.workbook["Metadata"]
        self.additional_indicators_worksheet = self.workbook["Additional Indicators"]

        last_results_row = self.results_worksheet.max_row + 1
        last_metadata_row = self.metadata_worksheet.max_row + 1
        last_additional_indicators_row = self.additional_indicators_worksheet.max_row + 1

        self.metadata_worksheet.cell(row=last_metadata_row, column=1, value="Set Aside")
        self.metadata_worksheet.cell(row=last_metadata_row, column=1, value="Set Aside").fill = Colors.LIGHT_BLUE_FILL.value
        self.metadata_worksheet.cell(row=last_metadata_row + 1, column=1, value="Hectares")
        self.metadata_worksheet.cell(row=last_metadata_row + 2, column=1, value="Is set aside")

        if self.module.is_start():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=2, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=2, value=self.module.is_set_aside_start)

        if self.module.is_with():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=3, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=3, value=self.module.is_set_aside_w)

        if self.module.is_without():
            self.metadata_worksheet.cell(row=last_metadata_row + 1, column=4, value=self.units)
            self.metadata_worksheet.cell(row=last_metadata_row + 2, column=4, value=self.module.is_set_aside_wo)

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class GrasslandReport(LandModuleReport):

    module: api_models.Grassland
    activity_report: BaseActivityReport = None

    def __post_init__(self):
        self.calculator = calculators.GrasslandCalculator(self.module)
        return super().__post_init__()


@dataclass
class OtherLandReport(LandModuleReport):

    module: api_models.OtherLand
    activity_report: BaseActivityReport = None

    def __post_init__(self):
        self.calculator = calculators.OtherLandCalculator(self.module)
        return super().__post_init__()


@dataclass
class CoastalWetlandReport(LandModuleReport):

    module: api_models.CoastalWetland
    activity_report: BaseActivityReport = None

    def __post_init__(self):
        self.calculator = calculators.CoastalWetlandCalculator(self.module)
        return super().__post_init__()


@dataclass
class FloodedRiceReport(LandModuleReport):

    module: api_models.FloodedRice
    activity_report: BaseActivityReport = None

    rice_cultivation_ch4: list[float] = None

    fire_n2o_source = (math_utils.ActivityTypes.STRAW_BURNING, math_utils.GasTypes.N2O)
    fire_ch4_source = (math_utils.ActivityTypes.STRAW_BURNING, math_utils.GasTypes.CH4)
    rice_cultivation_ch4_source = (math_utils.ActivityTypes.CH4_EMITTED_RICE, math_utils.GasTypes.CH4)

    def __post_init__(self):
        self.calculator = calculators.FloodedRiceCalculator(self.module)
        return super().__post_init__()

    def add_minor_seasons_results(self):
        minor_seasons = getattr(self.module, "submodules", [])

        for minor_season in minor_seasons.all():
            log.debug(f"Building report for minor season {minor_season.name}")

            minor_calculator = calculators.FloodedRiceCalculator(minor_season)
            minor_emission_set = []

            if self.module.is_with():
                minor_emission_set += minor_calculator.results_w.yearly_emissions_by_sector_by_gas
                if self.calculator.results_start_w is not None:
                    minor_emission_set += self.calculator.results_start_w.yearly_emissions_by_sector_by_gas

            if self.module.is_without():
                minor_emission_set += minor_calculator.results_wo.yearly_emissions_by_sector_by_gas
                if self.calculator.results_start_wo is not None:
                    minor_emission_set += self.calculator.results_start_wo.yearly_emissions_by_sector_by_gas

            self.biomass_co2 = [x + y for x, y in zip(self.biomass_co2, self.extract_emissions(minor_emission_set, self.biomass_co2_source[0], self.biomass_co2_source[1]))]
            self.soil_co2 = [x + y for x, y in zip(self.soil_co2, self.extract_emissions(minor_emission_set, self.soil_co2_source[0], self.soil_co2_source[1]))]
            self.soil_n2o = [x + y for x, y in zip(self.soil_n2o, self.extract_emissions(minor_emission_set, self.soil_n2o_source[0], self.soil_n2o_source[1]))]
            self.fire_n2o = [x + y for x, y in zip(self.fire_n2o, self.extract_emissions(minor_emission_set, self.fire_n2o_source[0], self.fire_n2o_source[1]))]
            self.fire_ch4 = [x + y for x, y in zip(self.fire_ch4, self.extract_emissions(minor_emission_set, self.fire_ch4_source[0], self.fire_ch4_source[1]))]
            self.rice_cultivation_ch4 = [x + y for x, y in zip(self.rice_cultivation_ch4, self.extract_emissions(minor_emission_set, self.rice_cultivation_ch4_source[0], self.rice_cultivation_ch4_source[1]))]

    def build_report(self):
        super().build_report()

        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.rice_cultivation_ch4 = self.extract_emissions(self.emissions_set, self.rice_cultivation_ch4_source[0], self.rice_cultivation_ch4_source[1])

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CH4 from rice cultivation")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.rice_cultivation_ch4[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class WaterbodyReport(BaseModuleReport):

    module: api_models.Waterbody

    waterbody_management_ch4: list[float] = None
    waterbody_management_ch4_source = (math_utils.ActivityTypes.COASTAL_WATERBODIES, math_utils.GasTypes.CH4)

    def __post_init__(self):
        self.calculator = calculators.WaterbodyCalculator(self.module)
        return super().__post_init__()

    def build_report(self):

        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.waterbody_management_ch4 = self.extract_emissions(self.emissions_set, self.waterbody_management_ch4_source[0], self.waterbody_management_ch4_source[1])

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CH4 from waterbody management")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.waterbody_management_ch4[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class AquacultureReport(BaseModuleReport):

    module: api_models.Aquaculture

    fish_n2o: list[float] = None
    electricity_co2_eq: list[float] = None

    fish_n2o_source = (math_utils.ActivityTypes.N20_FIELD, math_utils.GasTypes.N2O)
    electricity_co2_eq_source = (math_utils.ActivityTypes.ELECTRICITY, math_utils.GasTypes.CO2)

    def __post_init__(self):
        self.calculator = calculators.AquacultureCalculator(self.module)
        return super().__post_init__()

    def build_report(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.fish_n2o = self.extract_emissions(self.emissions_set, self.fish_n2o_source[0], self.fish_n2o_source[1])
        self.electricity_co2_eq = self.extract_emissions(self.emissions_set, self.electricity_co2_eq_source[0], self.electricity_co2_eq_source[1])

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="N2O from fish")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CO2-eq from electricity")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.fish_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.electricity_co2_eq[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class FisheryReport(BaseModuleReport):

    liquid_fuel_co2: list[float] = None
    liquid_fuel_n2o: list[float] = None
    liquid_fuel_ch4: list[float] = None
    refrigeration_hfc: list[float] = None
    electricity_co2_eq: list[float] = None

    liquid_fuel_co2_source = (math_utils.ActivityTypes.CATCH, math_utils.GasTypes.CO2)
    liquid_fuel_n2o_source = (math_utils.ActivityTypes.CATCH, math_utils.GasTypes.N2O)
    liquid_fuel_ch4_source = (math_utils.ActivityTypes.CATCH, math_utils.GasTypes.CH4)
    refrigeration_hfc_source = (math_utils.ActivityTypes.REFRIGERANT, math_utils.GasTypes.OTHER)
    electricity_co2_eq_source = (math_utils.ActivityTypes.ICE, math_utils.GasTypes.OTHER)

    def __post_init__(self):
        return super().__post_init__()

    def build_report(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.liquid_fuel_co2 = self.extract_emissions(self.emissions_set, self.liquid_fuel_co2_source[0], self.liquid_fuel_co2_source[1])
        self.liquid_fuel_n2o = self.extract_emissions(self.emissions_set, self.liquid_fuel_n2o_source[0], self.liquid_fuel_n2o_source[1])
        self.liquid_fuel_ch4 = self.extract_emissions(self.emissions_set, self.liquid_fuel_ch4_source[0], self.liquid_fuel_ch4_source[1])
        self.refrigeration_hfc = self.extract_emissions(self.emissions_set, self.refrigeration_hfc_source[0], self.refrigeration_hfc_source[1])
        self.electricity_co2_eq = self.extract_emissions(self.emissions_set, self.electricity_co2_eq_source[0], self.electricity_co2_eq_source[1])

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CO2 from liquid fuels consumption")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="N2O from liquid fuels consumption")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="CH4 from liquid fuels consumption")
        self.results_worksheet.cell(row=last_results_row + 4, column=1, value="HFC from refrigeration")
        self.results_worksheet.cell(row=last_results_row + 5, column=1, value="CO2-eq from electricity")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.liquid_fuel_co2[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.liquid_fuel_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.liquid_fuel_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 4, column=i + 2, value=self.refrigeration_hfc[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class SmallFisheryReport(FisheryReport):

    module: api_models.SmallFishery

    def __post_init__(self):
        self.calculator = calculators.SmallFisheryCalculator(self.module)
        return super().__post_init__()

    def build_report(self):
        return super().build_report()


@dataclass
class LargeFisheryReport(FisheryReport):

    module: api_models.LargeFishery

    def __post_init__(self):
        self.calculator = calculators.LargeFisheryCalculator(self.module)
        return super().__post_init__()

    def build_report(self):
        return super().build_report()


class LivestockReport(BaseModuleReport):

    module: api_models.Livestock

    enteric_fermentation_ch4: list[float] = None
    manure_management_other_than_prp_ch4: list[float] = None
    manure_management_other_than_prp_direct_n2o: list[float] = None
    manure_management_other_than_prp_leaching_indirect_n2o: list[float] = None
    manure_management_other_than_prp_volatilization_indirect_n2o: list[float] = None
    manure_management_prp_ch4: list[float] = None
    manure_management_prp_direct_n2o: list[float] = None
    manure_management_prp_leaching_indirect_n2o: list[float] = None
    manure_management_prp_volatilization_indirect_n2o: list[float] = None

    enteric_fermentation_ch4_source = (math_utils.ActivityTypes.METHANE_ENTERIC_FERMENTATION, math_utils.GasTypes.CH4)
    manure_management_other_than_prp_ch4_source = (math_utils.ActivityTypes.METHANE_MANURE_MANAGEMENT_SYSTEM, math_utils.GasTypes.CH4)
    manure_management_other_than_prp_direct_n2o_source = (math_utils.ActivityTypes.NITROUS_MANURE_MANAGEMENT_SYSTEM, math_utils.GasTypes.N2O)
    manure_management_other_than_prp_leaching_indirect_n2o_source = (math_utils.ActivityTypes.NITROUS_MANURE_MANAGEMENT_INDIRECT_LEACHING_SYSTEM, math_utils.GasTypes.N2O)
    manure_management_other_than_prp_volatilization_indirect_n2o_source = (math_utils.ActivityTypes.NITROUS_MANURE_MANAGEMENT_INDIRECT_VOLATILIZATION_SYSTEM, math_utils.GasTypes.N2O)
    manure_management_prp_ch4_source = (math_utils.ActivityTypes.METHANE_MANURE_MANAGEMENT_PRP, math_utils.GasTypes.CH4)
    manure_management_prp_direct_n2o_source = (math_utils.ActivityTypes.NITROUS_MANURE_MANAGEMENT_PRP, math_utils.GasTypes.N2O)
    manure_management_prp_leaching_indirect_n2o_source = (math_utils.ActivityTypes.NITROUS_MANURE_MANAGEMENT_INDIRECT_LEACHING_PRP, math_utils.GasTypes.N2O)
    manure_management_prp_volatilization_indirect_n2o_source = (math_utils.ActivityTypes.NITROUS_MANURE_MANAGEMENT_INDIRECT_VOLATILIZATION_PRP, math_utils.GasTypes.N2O)

    def __post_init__(self):
        self.calculator = calculators.LivestockCalculator(self.module)
        return super().__post_init__()

    def build_report(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.enteric_fermentation_ch4 = self.extract_emissions(self.emissions_set, self.enteric_fermentation_ch4_source[0], self.enteric_fermentation_ch4_source[1])
        self.manure_management_other_than_prp_ch4 = self.extract_emissions(self.emissions_set, self.manure_management_other_than_prp_ch4_source[0], self.manure_management_other_than_prp_ch4_source[1])
        self.manure_management_other_than_prp_direct_n2o = self.extract_emissions(self.emissions_set, self.manure_management_other_than_prp_direct_n2o_source[0], self.manure_management_other_than_prp_direct_n2o_source[1])
        self.manure_management_other_than_prp_leaching_indirect_n2o = self.extract_emissions(self.emissions_set, self.manure_management_other_than_prp_leaching_indirect_n2o_source[0], self.manure_management_other_than_prp_leaching_indirect_n2o_source[1])
        self.manure_management_other_than_prp_volatilization_indirect_n2o = self.extract_emissions(self.emissions_set, self.manure_management_other_than_prp_volatilization_indirect_n2o_source[0], self.manure_management_other_than_prp_volatilization_indirect_n2o_source[1])
        self.manure_management_prp_ch4 = self.extract_emissions(self.emissions_set, self.manure_management_prp_ch4_source[0], self.manure_management_prp_ch4_source[1])
        self.manure_management_prp_direct_n2o = self.extract_emissions(self.emissions_set, self.manure_management_prp_direct_n2o_source[0], self.manure_management_prp_direct_n2o_source[1])
        self.manure_management_prp_leaching_indirect_n2o = self.extract_emissions(self.emissions_set, self.manure_management_prp_leaching_indirect_n2o_source[0], self.manure_management_prp_leaching_indirect_n2o_source[1])
        self.manure_management_prp_volatilization_indirect_n2o = self.extract_emissions(self.emissions_set, self.manure_management_prp_volatilization_indirect_n2o_source[0], self.manure_management_prp_volatilization_indirect_n2o_source[1])

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CH4 from enteric fermentation")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CH4 from manure management other than PRP")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="Direct N2O from manure management other than PRP (direct)")
        self.results_worksheet.cell(row=last_results_row + 4, column=1, value="Indirect N2O from manure management other than PRP (leaching)")
        self.results_worksheet.cell(row=last_results_row + 5, column=1, value="Indirect N2O from manure management other than PRP (volatilization)")
        self.results_worksheet.cell(row=last_results_row + 6, column=1, value="CH4 from manure management PRP")
        self.results_worksheet.cell(row=last_results_row + 7, column=1, value="Direct N2O from manure management PRP (direct)")
        self.results_worksheet.cell(row=last_results_row + 8, column=1, value="Indirect N2O from manure management PRP (leaching)")
        self.results_worksheet.cell(row=last_results_row + 9, column=1, value="Indirect N2O from manure management PRP (volatilization)")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.enteric_fermentation_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.manure_management_other_than_prp_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.manure_management_other_than_prp_direct_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 4, column=i + 2, value=self.manure_management_other_than_prp_leaching_indirect_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 5, column=i + 2, value=self.manure_management_other_than_prp_volatilization_indirect_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 6, column=i + 2, value=self.manure_management_prp_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 7, column=i + 2, value=self.manure_management_prp_direct_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 8, column=i + 2, value=self.manure_management_prp_leaching_indirect_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 9, column=i + 2, value=self.manure_management_prp_volatilization_indirect_n2o[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class ForestManagementReport(LandModuleReport):

    module: api_models.ForestManagement

    hwp_co2: list[float] = None
    biomass_loss_co2: list[float] = None
    biomass_gain_co2: list[float] = None

    rotation_hwp_agb_co2: list[float] = None
    rotation_hwp_bgb_co2: list[float] = None
    rotation_agb_co2: list[float] = None
    rotation_agb_ch4: list[float] = None
    rotation_agb_n2o: list[float] = None
    rotation_bgb_co2: list[float] = None
    rotation_bgb_ch4: list[float] = None
    rotation_bgb_n2o: list[float] = None
    disturbance_agb_co2: list[float] = None
    disturbance_agb_ch4: list[float] = None
    disturbance_agb_n2o: list[float] = None
    disturbance_bgb_co2: list[float] = None
    disturbance_bgb_ch4: list[float] = None
    disturbance_bgb_n2o: list[float] = None
    logging_hwp_agb_co2: list[float] = None
    logging_hwp_bgb_co2: list[float] = None
    logging_agb_co2: list[float] = None
    logging_agb_ch4: list[float] = None
    logging_agb_n2o: list[float] = None
    logging_bgb_co2: list[float] = None
    logging_bgb_ch4: list[float] = None
    logging_bgb_n2o: list[float] = None
    degradation_agb_co2: list[float] = None
    degradation_bgb_co2: list[float] = None
    degradation_litter_co2: list[float] = None
    degradation_deadwood_co2: list[float] = None
    growth_agb_co2: list[float] = None
    growth_bgb_co2: list[float] = None
    litter_co2: list[float] = None
    deadwood_co2: list[float] = None

    hwp_rotation_logging_agb_co2_source = (math_utils.ActivityTypes.HWP_ROTATION_AGB, math_utils.GasTypes.CO2)
    hwp_rotation_logging_bgb_co2_source = (math_utils.ActivityTypes.HWP_ROTATION_BGB, math_utils.GasTypes.CO2)
    rotation_agb_n2o_source = (math_utils.ActivityTypes.ROTATION_AGB, math_utils.GasTypes.N2O)
    rotation_agb_ch4_source = (math_utils.ActivityTypes.ROTATION_AGB, math_utils.GasTypes.CH4)
    rotation_bgb_n2o_source = (math_utils.ActivityTypes.ROTATION_BGB, math_utils.GasTypes.N2O)
    rotation_bgb_ch4_source = (math_utils.ActivityTypes.ROTATION_BGB, math_utils.GasTypes.CH4)
    rotation_agb_co2_source = (math_utils.ActivityTypes.ROTATION_AGB, math_utils.GasTypes.CO2)
    rotation_bgb_co2_source = (math_utils.ActivityTypes.ROTATION_BGB, math_utils.GasTypes.CO2)
    disturbance_agb_co2_source = (math_utils.ActivityTypes.DISTURBANCE_AGB, math_utils.GasTypes.CO2)
    disturbance_bgb_co2_source = (math_utils.ActivityTypes.DISTURBANCE_BGB, math_utils.GasTypes.CO2)
    disturbance_fire_agb_n2o_source = (math_utils.ActivityTypes.DISTURBANCE_FIRE_AGB, math_utils.GasTypes.N2O)
    disturbance_fire_agb_ch4_source = (math_utils.ActivityTypes.DISTURBANCE_FIRE_AGB, math_utils.GasTypes.CH4)
    disturbance_fire_bgb_n2o_source = (math_utils.ActivityTypes.DISTURBANCE_FIRE_BGB, math_utils.GasTypes.N2O)
    disturbance_fire_bgb_ch4_source = (math_utils.ActivityTypes.DISTURBANCE_FIRE_BGB, math_utils.GasTypes.CH4)
    disturbance_fire_agb_co2_source = (math_utils.ActivityTypes.DISTURBANCE_FIRE_AGB, math_utils.GasTypes.CO2)
    disturbance_fire_bgb_co2_source = (math_utils.ActivityTypes.DISTURBANCE_FIRE_BGB, math_utils.GasTypes.CO2)
    hwp_logging_agb_co2_source = (math_utils.ActivityTypes.HWP_LOGGING_AGB, math_utils.GasTypes.CO2)
    hwp_logging_bgb_co2_source = (math_utils.ActivityTypes.HWP_LOGGING_BGB, math_utils.GasTypes.CO2)
    logging_agb_n2o_source = (math_utils.ActivityTypes.LOGGING_AGB, math_utils.GasTypes.N2O)
    logging_agb_ch4_source = (math_utils.ActivityTypes.LOGGING_AGB, math_utils.GasTypes.CH4)
    logging_bgb_n2o_source = (math_utils.ActivityTypes.LOGGING_BGB, math_utils.GasTypes.N2O)
    logging_bgb_ch4_source = (math_utils.ActivityTypes.LOGGING_BGB, math_utils.GasTypes.CH4)
    logging_agb_co2_source = (math_utils.ActivityTypes.LOGGING_AGB, math_utils.GasTypes.CO2)
    logging_bgb_co2_source = (math_utils.ActivityTypes.LOGGING_BGB, math_utils.GasTypes.CO2)
    degradation_agb_co2_source = (math_utils.ActivityTypes.DEGRADATION_AGB, math_utils.GasTypes.CO2)
    degradation_bgb_co2_source = (math_utils.ActivityTypes.DEGRADATION_BGB, math_utils.GasTypes.CO2)
    growth_agb_co2_source = (math_utils.ActivityTypes.AGB_GROWTH, math_utils.GasTypes.CO2)
    growth_bgb_co2_source = (math_utils.ActivityTypes.BGB_GROWTH, math_utils.GasTypes.CO2)
    degradation_litter_co2_source = (math_utils.ActivityTypes.DEGRADATION_LITTER, math_utils.GasTypes.CO2)
    litter_co2_source = (math_utils.ActivityTypes.LITTER, math_utils.GasTypes.CO2)
    degradation_deadwood_co2_source = (math_utils.ActivityTypes.DEGRADATION_DEADWOOD, math_utils.GasTypes.CO2)
    deadwood_co2_source = (math_utils.ActivityTypes.DEADWOOD, math_utils.GasTypes.CO2)

    def __post_init__(self):
        self.calculator = calculators.ForestManagementCalculator(self.module)
        return super().__post_init__()

    def build_report(self):
        super().build_report()

        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.rotation_hwp_agb_co2 = self.extract_emissions(self.emissions_set, self.hwp_rotation_logging_agb_co2_source[0], self.hwp_rotation_logging_agb_co2_source[1])
        self.rotation_hwp_bgb_co2 = self.extract_emissions(self.emissions_set, self.hwp_rotation_logging_bgb_co2_source[0], self.hwp_rotation_logging_bgb_co2_source[1])
        self.rotation_agb_n2o = self.extract_emissions(self.emissions_set, self.rotation_agb_n2o_source[0], self.rotation_agb_n2o_source[1])
        self.rotation_agb_ch4 = self.extract_emissions(self.emissions_set, self.rotation_agb_ch4_source[0], self.rotation_agb_ch4_source[1])
        self.rotation_bgb_n2o = self.extract_emissions(self.emissions_set, self.rotation_bgb_n2o_source[0], self.rotation_bgb_n2o_source[1])
        self.rotation_bgb_ch4 = self.extract_emissions(self.emissions_set, self.rotation_bgb_ch4_source[0], self.rotation_bgb_ch4_source[1])
        self.rotation_agb_co2 = self.extract_emissions(self.emissions_set, self.rotation_agb_co2_source[0], self.rotation_agb_co2_source[1])
        self.rotation_bgb_co2 = self.extract_emissions(self.emissions_set, self.rotation_bgb_co2_source[0], self.rotation_bgb_co2_source[1])
        self.disturbance_agb_co2 = self.extract_emissions(self.emissions_set, self.disturbance_agb_co2_source[0], self.disturbance_agb_co2_source[1])
        self.disturbance_bgb_co2 = self.extract_emissions(self.emissions_set, self.disturbance_bgb_co2_source[0], self.disturbance_bgb_co2_source[1])
        self.disturbance_agb_n2o = self.extract_emissions(self.emissions_set, self.disturbance_fire_agb_n2o_source[0], self.disturbance_fire_agb_n2o_source[1])
        self.disturbance_agb_ch4 = self.extract_emissions(self.emissions_set, self.disturbance_fire_agb_ch4_source[0], self.disturbance_fire_agb_ch4_source[1])
        self.disturbance_bgb_n2o = self.extract_emissions(self.emissions_set, self.disturbance_fire_bgb_n2o_source[0], self.disturbance_fire_bgb_n2o_source[1])
        self.disturbance_bgb_ch4 = self.extract_emissions(self.emissions_set, self.disturbance_fire_bgb_ch4_source[0], self.disturbance_fire_bgb_ch4_source[1])
        self.logging_hwp_agb_co2 = self.extract_emissions(self.emissions_set, self.hwp_logging_agb_co2_source[0], self.hwp_logging_agb_co2_source[1])
        self.logging_hwp_bgb_co2 = self.extract_emissions(self.emissions_set, self.hwp_logging_bgb_co2_source[0], self.hwp_logging_bgb_co2_source[1])
        self.logging_agb_n2o = self.extract_emissions(self.emissions_set, self.logging_agb_n2o_source[0], self.logging_agb_n2o_source[1])
        self.logging_agb_ch4 = self.extract_emissions(self.emissions_set, self.logging_agb_ch4_source[0], self.logging_agb_ch4_source[1])
        self.logging_bgb_n2o = self.extract_emissions(self.emissions_set, self.logging_bgb_n2o_source[0], self.logging_bgb_n2o_source[1])
        self.logging_bgb_ch4 = self.extract_emissions(self.emissions_set, self.logging_bgb_ch4_source[0], self.logging_bgb_ch4_source[1])
        self.logging_agb_co2 = self.extract_emissions(self.emissions_set, self.logging_agb_co2_source[0], self.logging_agb_co2_source[1])
        self.logging_bgb_co2 = self.extract_emissions(self.emissions_set, self.logging_bgb_co2_source[0], self.logging_bgb_co2_source[1])
        self.degradation_agb_co2 = self.extract_emissions(self.emissions_set, self.degradation_agb_co2_source[0], self.degradation_agb_co2_source[1])
        self.degradation_bgb_co2 = self.extract_emissions(self.emissions_set, self.degradation_bgb_co2_source[0], self.degradation_bgb_co2_source[1])
        self.growth_agb_co2 = self.extract_emissions(self.emissions_set, self.growth_agb_co2_source[0], self.growth_agb_co2_source[1])
        self.growth_bgb_co2 = self.extract_emissions(self.emissions_set, self.growth_bgb_co2_source[0], self.growth_bgb_co2_source[1])
        self.litter_co2 = self.extract_emissions(self.emissions_set, self.litter_co2_source[0], self.litter_co2_source[1])
        self.deadwood_co2 = self.extract_emissions(self.emissions_set, self.deadwood_co2_source[0], self.deadwood_co2_source[1])
        self.degradation_litter_co2 = self.extract_emissions(self.emissions_set, self.degradation_litter_co2_source[0], self.degradation_litter_co2_source[1])
        self.degradation_deadwood_co2 = self.extract_emissions(self.emissions_set, self.degradation_deadwood_co2_source[0], self.degradation_deadwood_co2_source[1])

        self.hwp_co2 = list(map(sum, zip(self.rotation_hwp_agb_co2, self.rotation_hwp_bgb_co2, self.logging_hwp_agb_co2, self.logging_hwp_bgb_co2)))
        self.fire_n2o = list(map(sum, zip(self.rotation_agb_n2o, self.rotation_bgb_n2o, self.disturbance_agb_n2o, self.disturbance_bgb_n2o, self.logging_agb_n2o, self.logging_bgb_n2o)))
        self.fire_ch4 = list(map(sum, zip(self.rotation_agb_ch4, self.rotation_bgb_ch4, self.disturbance_agb_ch4, self.disturbance_bgb_ch4, self.logging_agb_ch4, self.logging_bgb_ch4)))
        self.biomass_loss_co2 = list(map(sum, zip(self.rotation_agb_co2, self.rotation_bgb_co2, self.disturbance_agb_co2, self.disturbance_bgb_co2, self.logging_agb_co2, self.logging_bgb_co2, self.degradation_agb_co2, self.degradation_bgb_co2, self.degradation_litter_co2, self.degradation_deadwood_co2)))
        self.biomass_gain_co2 = list(map(sum, zip(self.growth_agb_co2, self.growth_bgb_co2, self.litter_co2, self.deadwood_co2)))

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CO2 from HWP (rotation and logging)")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CO2 from biomass loss")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="CO2 from biomass gain")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=self.fire_n2o_row_index, column=i + 2, value=self.fire_n2o[i])
            self.results_worksheet.cell(row=self.fire_ch4_row_index, column=i + 2, value=self.fire_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.hwp_co2[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.biomass_loss_co2[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.biomass_gain_co2[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class EnergyReport(BaseModuleReport):
    module: api_models.Energy

    electricity_co2_eq: list[float] = None
    liquid_fuel_co2: list[float] = None
    liquid_fuel_ch4: list[float] = None
    liquid_fuel_n2o: list[float] = None
    solid_fuel_co2: list[float] = None
    solid_fuel_ch4: list[float] = None
    solid_fuel_n2o: list[float] = None

    electricity_co2_eq_source = (math_utils.ActivityTypes.ELECTRICITY, math_utils.GasTypes.CO2)
    liquid_fuel_co2_source = (math_utils.ActivityTypes.FUEL, math_utils.GasTypes.CO2)
    liquid_fuel_ch4_source = (math_utils.ActivityTypes.FUEL, math_utils.GasTypes.CH4)
    liquid_fuel_n2o_source = (math_utils.ActivityTypes.FUEL, math_utils.GasTypes.N2O)
    solid_fuel_co2_source = (math_utils.ActivityTypes.FUEL, math_utils.GasTypes.CO2)
    solid_fuel_ch4_source = (math_utils.ActivityTypes.FUEL, math_utils.GasTypes.CH4)
    solid_fuel_n2o_source = (math_utils.ActivityTypes.FUEL, math_utils.GasTypes.N2O)

    def __post_init__(self):
        self.calculator = calculators.EnergyCalculator(self.module)
        return super().__post_init__()

    def add_submodules_results(self):

        submodules: list[api_models.Submodule] = self.module.submodules

        for submodule in submodules:
            CalculatorClass = calculators.ElectricityCalculator if isinstance(submodule, api_models.Electricity) else calculators.FuelCalculator
            submodule: api_models.Electricity | api_models.Fuel

            calculator = CalculatorClass(submodule)
            from api.calculators import Result

            try:
                submodule_emission_set = Result(*calculator.calculate()).balance.yearly_emissions_by_sector_by_gas
            except Exception as e:
                log.error(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")
                raise NotReadyError(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")

            self.electricity_co2_eq = list(map(sum, zip(self.electricity_co2_eq, self.extract_emissions(submodule_emission_set, self.electricity_co2_eq_source[0], self.electricity_co2_eq_source[1]))))

            if isinstance(submodule, api_models.Fuel):
                if "solid" in submodule.fuel_type.macro_fuel_type.name.casefold():
                    self.solid_fuel_co2 = list(map(sum, zip(self.solid_fuel_co2, self.extract_emissions(submodule_emission_set, self.solid_fuel_co2_source[0], self.solid_fuel_co2_source[1]))))
                    self.solid_fuel_ch4 = list(map(sum, zip(self.solid_fuel_ch4, self.extract_emissions(submodule_emission_set, self.solid_fuel_ch4_source[0], self.solid_fuel_ch4_source[1]))))
                    self.solid_fuel_n2o = list(map(sum, zip(self.solid_fuel_n2o, self.extract_emissions(submodule_emission_set, self.solid_fuel_n2o_source[0], self.solid_fuel_n2o_source[1]))))
                elif "liquid" in submodule.fuel_type.macro_fuel_type.name.casefold():
                    self.liquid_fuel_co2 = list(map(sum, zip(self.liquid_fuel_co2, self.extract_emissions(submodule_emission_set, self.liquid_fuel_co2_source[0], self.liquid_fuel_co2_source[1]))))
                    self.liquid_fuel_ch4 = list(map(sum, zip(self.liquid_fuel_ch4, self.extract_emissions(submodule_emission_set, self.liquid_fuel_ch4_source[0], self.liquid_fuel_ch4_source[1]))))
                    self.liquid_fuel_n2o = list(map(sum, zip(self.liquid_fuel_n2o, self.extract_emissions(submodule_emission_set, self.liquid_fuel_n2o_source[0], self.liquid_fuel_n2o_source[1]))))

    def build_report(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.electricity_co2_eq = np.zeros(self.last_year_of_accounting - self.start_year_of_activities)
        self.liquid_fuel_co2 = np.zeros(self.last_year_of_accounting - self.start_year_of_activities)
        self.liquid_fuel_ch4 = np.zeros(self.last_year_of_accounting - self.start_year_of_activities)
        self.liquid_fuel_n2o = np.zeros(self.last_year_of_accounting - self.start_year_of_activities)
        self.solid_fuel_co2 = np.zeros(self.last_year_of_accounting - self.start_year_of_activities)
        self.solid_fuel_ch4 = np.zeros(self.last_year_of_accounting - self.start_year_of_activities)
        self.solid_fuel_n2o = np.zeros(self.last_year_of_accounting - self.start_year_of_activities)

        self.add_submodules_results()

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CO2-eq from electricity")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CO2 from liquid fuels")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="CH4 from liquid fuels")
        self.results_worksheet.cell(row=last_results_row + 4, column=1, value="N2O from liquid fuels")
        self.results_worksheet.cell(row=last_results_row + 5, column=1, value="CO2 from solid fuels")
        self.results_worksheet.cell(row=last_results_row + 6, column=1, value="CH4 from solid fuels")
        self.results_worksheet.cell(row=last_results_row + 7, column=1, value="N2O from solid fuels")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.electricity_co2_eq[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.liquid_fuel_co2[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.liquid_fuel_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 4, column=i + 2, value=self.liquid_fuel_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 5, column=i + 2, value=self.solid_fuel_co2[i])
            self.results_worksheet.cell(row=last_results_row + 6, column=i + 2, value=self.solid_fuel_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 7, column=i + 2, value=self.solid_fuel_n2o[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class InputReport(BaseModuleReport):

    module: api_models.Input

    inputs_co2: list[float] = None
    inputs_n2o: list[float] = None
    inputs_co2_eq: list[float] = None
    feed_co2_eq: list[float] = None

    inputs_co2_source = (math_utils.ActivityTypes.CO2_FIELD, math_utils.GasTypes.CO2)
    inputs_n2o_source = (math_utils.ActivityTypes.N20_FIELD, math_utils.GasTypes.N2O)
    inputs_co2_eq_source = (math_utils.ActivityTypes.CO2_EQUIVALENT_VC, math_utils.GasTypes.CO2)
    feed_co2_eq_source = (math_utils.ActivityTypes.CO2_EQUIVALENT_VC, math_utils.GasTypes.CO2)

    def __post_init__(self):
        self.calculator = calculators.InputCalculator(self.module)
        return super().__post_init__()

    def add_submodules_results(self):
        submodules: list[api_models.Submodule] = self.module.submodules

        for submodule in submodules:
            CalculatorClass = calculators.InputEntryCalculator
            submodule: api_models.InputEntry
            submodule_emission_set = []

            calculator = CalculatorClass(submodule)
            from api.calculators import Result

            try:
                submodule_emission_set = Result(*calculator.calculate()).balance.yearly_emissions_by_sector_by_gas
            except Exception as e:
                log.error(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")
                raise NotReadyError(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")

            self.inputs_co2 = list(map(sum, zip(self.inputs_co2, self.extract_emissions(submodule_emission_set, self.inputs_co2_source[0], self.inputs_co2_source[1]))))
            self.inputs_n2o = list(map(sum, zip(self.inputs_n2o, self.extract_emissions(submodule_emission_set, self.inputs_n2o_source[0], self.inputs_n2o_source[1]))))

            if "feed" in submodule.input_type.macro_input_type.name.casefold():
                self.feed_co2_eq = list(map(sum, zip(self.feed_co2_eq, self.extract_emissions(submodule_emission_set, self.feed_co2_eq_source[0], self.feed_co2_eq_source[1]))))
            else:
                self.inputs_co2_eq = list(map(sum, zip(self.inputs_co2_eq, self.extract_emissions(submodule_emission_set, self.inputs_co2_eq_source[0], self.inputs_co2_eq_source[1]))))

    def build_report(self):
        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        """
        # BUG: Calculator aggregates emissions from all submodules, but the report should only show the emissions from the main module and then add the emissions from the submodules separately
        Setting the emissions to zero to avoid double counting might be skipping some emissions from the main module (also happening in Input, Energy, ForestManageent, AnnualCropland and PerennialCropland reports)
        Update: this must be evaluated on a module-by-module basis. For parent modules that have no calculations attached to them, it's easier to just set the emissions to zero. For others, like ForestManagement and the Croplands, more testing is needed
        """

        self.inputs_co2 = np.zeros(self.duration)
        self.inputs_n2o = np.zeros(self.duration)
        self.inputs_co2_eq = np.zeros(self.duration)
        self.feed_co2_eq = np.zeros(self.duration)

        self.add_submodules_results()

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="N2O from inputs (field level)")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CO2 from inputs (field level)")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="CO2-eq from inputs (production, transportation and storage)")
        self.results_worksheet.cell(row=last_results_row + 4, column=1, value="CO2-eq from feed")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.inputs_co2[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.inputs_n2o[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.inputs_co2_eq[i])
            self.results_worksheet.cell(row=last_results_row + 4, column=i + 2, value=self.feed_co2_eq[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


@dataclass
class IrrigationReport(BaseModuleReport):

    module: api_models.Irrigation

    other_infrastructure_co2_eq: list[float] = None
    liquid_fuel_or_electricity_co2: list[float] = None
    liquid_fuel_or_electricity_ch4: list[float] = None
    liquid_fuel_or_electricity_n2o: list[float] = None

    other_infrastructure_co2_eq_source = (math_utils.ActivityTypes.NEW_IRRIGATION, math_utils.GasTypes.CO2)
    liquid_fuel_or_electricity_co2_source = (math_utils.ActivityTypes.IRRIGATION_OPERATIONAL, math_utils.GasTypes.CO2)
    liquid_fuel_or_electricity_ch4_source = (math_utils.ActivityTypes.IRRIGATION_OPERATIONAL, math_utils.GasTypes.CH4)
    liquid_fuel_or_electricity_n2o_source = (math_utils.ActivityTypes.IRRIGATION_OPERATIONAL, math_utils.GasTypes.N2O)

    def __post_init__(self):
        self.calculator = calculators.IrrigationCalculator(self.module)
        return super().__post_init__()

    def add_submodules_results(self):
        submodules: list[api_models.Submodule] = self.module.submodules

        self.other_infrastructure_co2_eq = np.zeros(self.duration)
        self.liquid_fuel_or_electricity_co2 = np.zeros(self.duration)
        self.liquid_fuel_or_electricity_ch4 = np.zeros(self.duration)
        self.liquid_fuel_or_electricity_n2o = np.zeros(self.duration)

        for submodule in submodules:
            submodules_emission_set = []
            CalculatorClass = calculators.IrrigationPhaseCalculator if isinstance(submodule, api_models.IrrigationPhase) else calculators.IrrigationSystemCalculator
            submodule: api_models.IrrigationPhase | api_models.IrrigationSystem

            calculator = CalculatorClass(submodule)
            from api.calculators import Result

            try:
                submodules_emission_set = Result(*calculator.calculate()).balance.yearly_emissions_by_sector_by_gas
            except Exception as e:
                log.error(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")
                raise NotReadyError(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")

            self.other_infrastructure_co2_eq = self.extract_emissions(submodules_emission_set, self.other_infrastructure_co2_eq_source[0], self.other_infrastructure_co2_eq_source[1])
            self.liquid_fuel_or_electricity_co2 = self.extract_emissions(submodules_emission_set, self.liquid_fuel_or_electricity_co2_source[0], self.liquid_fuel_or_electricity_co2_source[1])
            self.liquid_fuel_or_electricity_ch4 = self.extract_emissions(submodules_emission_set, self.liquid_fuel_or_electricity_ch4_source[0], self.liquid_fuel_or_electricity_ch4_source[1])
            self.liquid_fuel_or_electricity_n2o = self.extract_emissions(submodules_emission_set, self.liquid_fuel_or_electricity_n2o_source[0], self.liquid_fuel_or_electricity_n2o_source[1])

    def build_report(self):
        super().build_report()

        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.add_submodules_results()

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CO2-eq from other infrastructure")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CO2 from liquid fuel or electricity")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="CH4 from liquid fuel or electricity")
        self.results_worksheet.cell(row=last_results_row + 4, column=1, value="N2O from liquid fuel or electricity")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.other_infrastructure_co2_eq[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.liquid_fuel_or_electricity_co2[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.liquid_fuel_or_electricity_ch4[i])
            self.results_worksheet.cell(row=last_results_row + 4, column=i + 2, value=self.liquid_fuel_or_electricity_n2o[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)


# @dataclass
class SettlementReport(LandModuleReport):

    module: api_models.Settlement

    buildings_co2_eq: list[float] = None
    roads_co2_eq: list[float] = None
    other_infrastructure_co2_eq: list[float] = None

    buildings_co2_eq_source = (math_utils.ActivityTypes.ROADS, math_utils.GasTypes.CO2)
    roads_co2_eq_source = (math_utils.ActivityTypes.ROADS, math_utils.GasTypes.CO2)
    infrastructure_co2_eq_source = (math_utils.ActivityTypes.ROADS, math_utils.GasTypes.CO2)

    def __post_init__(self):
        self.calculator = calculators.SettlementCalculator(self.module)
        self.module_endpoint = "settlements"
        return super().__post_init__()

    def add_submodules_results(self):
        submodules: list[api_models.Submodule] = self.module.submodules

        for submodule in submodules:
            if isinstance(submodule, api_models.OtherInfrastructure):
                continue

            submodules_emission_set = submodule.cached_results_by_activity_by_gas["balance"] if submodule.cached_results_by_activity_by_gas is not None else []
            CalculatorClass = calculators.RoadCalculator if isinstance(submodule, api_models.Road) else calculators.BuildingCalculator
            submodule: api_models.Road | api_models.Building

            calculator = CalculatorClass(submodule)
            from api.calculators import Result

            try:
                submodules_emission_set = Result(*calculator.calculate()).balance.yearly_emissions_by_sector_by_gas
            except Exception as e:
                log.error(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")
                raise NotReadyError(f"Cannot calculate emissions for submodule {submodule.module_type.name} in activity {submodule.parent.activity.name}: {e}")

            if isinstance(submodule, api_models.Road):
                self.roads_co2_eq = list(map(sum, zip(self.roads_co2_eq, self.extract_emissions(submodules_emission_set, self.roads_co2_eq_source[0], self.roads_co2_eq_source[1]))))
            elif isinstance(submodule, api_models.Building):
                self.buildings_co2_eq = list(map(sum, zip(self.buildings_co2_eq, self.extract_emissions(submodules_emission_set, self.buildings_co2_eq_source[0], self.buildings_co2_eq_source[1]))))

    def build_report(self):
        super().build_report()

        self.workbook = self.activity_report.project_report.excel_manager.get_workbook()
        self.results_worksheet = self.workbook["Results"]

        last_results_row = self.results_worksheet.max_row

        self.buildings_co2_eq = np.zeros(self.duration)
        self.roads_co2_eq = np.zeros(self.duration)
        self.other_infrastructure_co2_eq = np.zeros(self.duration)

        self.add_submodules_results()

        self.results_worksheet.cell(row=last_results_row + 1, column=1, value="CO2-eq from buildings")
        self.results_worksheet.cell(row=last_results_row + 2, column=1, value="CO2-eq from roads")
        self.results_worksheet.cell(row=last_results_row + 3, column=1, value="CO2-eq from other infrastructure")

        for i, year in enumerate(range(self.start_year_of_activities, self.last_year_of_accounting)):
            self.results_worksheet.cell(row=last_results_row + 1, column=i + 2, value=self.buildings_co2_eq[i])
            self.results_worksheet.cell(row=last_results_row + 2, column=i + 2, value=self.roads_co2_eq[i])
            self.results_worksheet.cell(row=last_results_row + 3, column=i + 2, value=self.other_infrastructure_co2_eq[i])

        self.activity_report.project_report.excel_manager.save_workbook(self.workbook)
