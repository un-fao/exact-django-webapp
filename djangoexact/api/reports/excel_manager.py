"""ExcelFileManager – handles in-memory Excel workbook lifecycle."""

from __future__ import annotations

import os
import tempfile
from datetime import datetime
from io import BytesIO

import openpyxl as pxl


class ExcelFileManager:
    def __init__(self):
        self._workbook: pxl.Workbook = self._create_initial_workbook()

    def _create_initial_workbook(self) -> pxl.Workbook:
        """Create the skeleton openpyxl workbook with required worksheets."""
        wb = pxl.Workbook()
        wb.active.title = "Results"
        wb.create_sheet("Metadata")
        ai_ws = wb.create_sheet("Additional Indicators")
        wb.create_sheet("Shadow Price of Carbon")
        wb.create_sheet("Inventory")
        ai_ws.sheet_state = "hidden"
        return wb

    def get_workbook(self) -> pxl.Workbook:
        """Return the live in-memory workbook (no serialization/deserialization)."""
        return self._workbook

    def save_workbook(self, workbook: pxl.Workbook) -> None:
        """No-op: workbook is kept in memory; call finalize() once at the end."""
        # Keep the reference in sync in case the caller passes it back
        self._workbook = workbook

    def finalize(self) -> bytes:
        """Serialize the workbook to bytes and optionally save to disk."""
        buf = BytesIO()
        self._workbook.save(buf)
        buf.seek(0)
        data = buf.getvalue()

        if os.environ.get("EXACT_SAVE_REPORTS_TO_FILE"):
            reports_dir = os.path.join(tempfile.gettempdir(), "reports")
            os.makedirs(reports_dir, exist_ok=True)
            filename = f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            filepath = os.path.join(reports_dir, filename)
            with open(filepath, "wb") as f:
                f.write(data)
            self.saved_report_path = filepath

        return data

    def get_excel_bytes(self) -> bytes:
        """Serialize and return the workbook as bytes."""
        return self.finalize()
