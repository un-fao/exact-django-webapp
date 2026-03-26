"""ExcelRenderer – renders a ProjectResult into Excel bytes.

All cell-writing logic lives here; no calculation is performed.
"""
from __future__ import annotations

import logging as log
from datetime import datetime
from itertools import zip_longest

import numpy as np
import openpyxl as pxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .constants import (
    DATA_COL_START,
    INV_COL_ACTIVITY,
    INV_COL_GAS_TYPE,
    INV_COL_IPCC_CATEGORY,
    INV_COL_MODULE,
    INV_COL_VALUE,
    META_COL_COMMENT,
    META_COL_LABEL,
    META_COL_START,
    META_COL_TIER2,
    META_COL_WITH,
    META_COL_WITHOUT,
    ROW_BIOMASS_CO2,
    ROW_CH4,
    ROW_CUMULATIVE_BALANCE,
    ROW_CUMULATIVE_HECTARES,
    ROW_CUMULATIVE_HEADS,
    ROW_CUMULATIVE_CATCH,
    ROW_N2O,
    ROW_OTHER_CO2,
    ROW_OTHER_GHGS,
    ROW_SOIL_CO2,
    ROW_YEARLY_BALANCE,
    SPC_NOMINAL_ROW,
    SPC_NPV_START_ROW,
    SPC_ROW_HEADER,
    SPC_ROW_WITH,
    SPC_ROW_WITH_HIGH,
    SPC_ROW_WITH_LOW,
    SPC_ROW_WO,
    SPC_ROW_WO_HIGH,
    SPC_ROW_WO_LOW,
    SPC_ROW_YEAR,
    WS_ADDITIONAL_INDICATORS,
    WS_INVENTORY,
    WS_METADATA,
    WS_RESULTS,
    WS_SHADOW_PRICE,
)
from .data_types import ActivityResult, ModuleResult, ProjectResult
from .excel_manager import ExcelFileManager

_ORANGE = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")
_BLUE = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
_BEIGE = PatternFill(start_color="ddd9c4", end_color="ddd9c4", fill_type="solid")
_RED = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")

_FILL_MAP = {
    "LIGHT_ORANGE": _ORANGE,
    "LIGHT_BLUE": _BLUE,
    "LIGHT_BEIGE": _BEIGE,
    "LIGHT_RED": _RED,
}


class ExcelRenderer:
    """Renders a ProjectResult into Excel bytes."""

    def __init__(self, result: ProjectResult) -> None:
        self.result = result
        self._manager = ExcelFileManager()

    def render(self) -> bytes:
        wb = self._manager.get_workbook()
        self._write_skeleton(wb)
        self._manager.save_workbook(wb)

        for activity_result in self.result.activity_results:
            wb = self._manager.get_workbook()
            self._write_activity(wb, activity_result)
            self._manager.save_workbook(wb)

        wb = self._manager.get_workbook()
        self._write_results_summary(wb)
        self._write_shadow_price(wb)
        self._write_inventory(wb)
        self._manager.save_workbook(wb)

        return self._manager.get_excel_bytes()

    # -------------------------------------------------------------------------
    # Skeleton – project-level headers
    # -------------------------------------------------------------------------

    def _write_skeleton(self, wb: pxl.Workbook) -> None:
        r = self.result
        proj = r.project
        ws_res = wb[WS_RESULTS]
        ws_meta = wb[WS_METADATA]
        ws_ai = wb[WS_ADDITIONAL_INDICATORS]
        ws_inv = wb[WS_INVENTORY]

        # Results headers
        ws_res.cell(row=1, column=1, value="Activity and GHGs / Years")
        ws_res.cell(row=2, column=1, value="Cumulative balance in Tco2-eq")
        ws_res.cell(row=3, column=1, value="Yearly balance in Tco2-eq")
        ws_res.cell(row=4, column=1, value="CO2 in biomass")
        ws_res.cell(row=5, column=1, value="CO2 in soils")
        ws_res.cell(row=6, column=1, value="Other CO2")
        ws_res.cell(row=7, column=1, value="CH4")
        ws_res.cell(row=8, column=1, value="N20")
        ws_res.cell(row=9, column=1, value="Other GHGs")
        ws_res.cell(row=10, column=1, value="Cumulative Hectares Impacted").fill = _ORANGE
        ws_res.cell(row=11, column=1, value="Cumulative Heads Impacted").fill = _ORANGE
        ws_res.cell(row=12, column=1, value="Cumulative Catch Impacted").fill = _ORANGE

        for i, year in enumerate(range(r.start_year, r.last_year)):
            ws_res.cell(row=1, column=i + DATA_COL_START, value=year)

        # Metadata headers
        ws_meta.cell(row=1, column=1, value="Report download date")
        ws_meta.cell(row=1, column=5, value="Tier 2")
        ws_meta.cell(row=1, column=6, value="Comments")
        ws_meta.cell(row=2, column=1, value="Project name")
        ws_meta.cell(row=3, column=1, value="Status")
        ws_meta.cell(row=4, column=1, value="Country")
        ws_meta.cell(row=5, column=1, value="Climate")
        ws_meta.cell(row=6, column=1, value="Moisture")
        ws_meta.cell(row=7, column=1, value="Soil type")
        ws_meta.cell(row=8, column=1, value="Implementation phase (years)")
        ws_meta.cell(row=9, column=1, value="Capitalization phase (years)")
        ws_meta.cell(row=10, column=1, value="Total duration of accounting (years)")
        ws_meta.cell(row=11, column=1, value="Global warming potential")
        ws_meta.cell(row=12, column=1, value="CO2")
        ws_meta.cell(row=13, column=1, value="CH4")
        ws_meta.cell(row=14, column=1, value="N2O")
        ws_meta.cell(row=15, column=1, value="Fossil CH4")

        ws_meta.cell(row=1, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws_meta.cell(row=2, column=2, value=str(proj.name))
        ws_meta.cell(row=3, column=2, value=proj.status.name if proj.status else "N/A")
        ws_meta.cell(row=4, column=2, value=proj.country.name)
        ws_meta.cell(row=5, column=2, value=proj.climate.name)
        ws_meta.cell(row=6, column=2, value=proj.moisture.name)
        ws_meta.cell(row=7, column=2, value=proj.soil_type.name)
        ws_meta.cell(row=8, column=2, value=proj.implementation_years)
        ws_meta.cell(row=9, column=2, value=proj.capitalization_years)
        ws_meta.cell(row=10, column=2, value=r.duration)
        ws_meta.cell(row=11, column=2, value=proj.gwp.name)
        ws_meta.cell(row=12, column=2, value=proj.gwp.co2)
        ws_meta.cell(row=13, column=2, value=proj.gwp.ch4)
        ws_meta.cell(row=14, column=2, value=proj.gwp.n2o)
        ws_meta.cell(row=15, column=2, value=proj.gwp.ch4_fossil)
        ws_meta.cell(row=16, column=1, value="")
        ws_meta.cell(row=17, column=1, value="")
        ws_meta.cell(row=18, column=2, value="START")
        ws_meta.cell(row=18, column=3, value="WITH")
        ws_meta.cell(row=18, column=4, value="WITHOUT")

        # Additional Indicators headers
        ws_ai.cell(row=1, column=1, value="Activity and GHGs / Years")
        for i, year in enumerate(range(r.start_year, r.last_year)):
            ws_ai.cell(row=1, column=i + DATA_COL_START, value=year)

        # Inventory headers
        for col, header in enumerate(
            ["Activity", "Module", "IPCC Category", "Gas Type", "Value (tCO2-eq)"], 1
        ):
            ws_inv.cell(row=1, column=col, value=header).fill = _ORANGE

    # -------------------------------------------------------------------------
    # Activity writing
    # -------------------------------------------------------------------------

    def _write_activity(self, wb: pxl.Workbook, ar: ActivityResult) -> None:
        ws_res = wb[WS_RESULTS]
        ws_meta = wb[WS_METADATA]
        ws_ai = wb[WS_ADDITIONAL_INDICATORS]
        dur = self.result.duration

        # Results: activity title row (values filled after modules)
        activity_row = ws_res.max_row + 1
        ws_res.cell(row=activity_row, column=1, value=str(ar.title)).fill = _ORANGE

        # Metadata: activity title + T2 overrides
        meta_row = ws_meta.max_row + 1
        ws_meta.cell(row=meta_row, column=1, value=str(ar.title)).fill = _BLUE
        for t2 in ar.t2_overrides:
            meta_row += 1
            ws_meta.cell(row=meta_row, column=META_COL_LABEL, value=t2.label)
            ws_meta.cell(row=meta_row, column=META_COL_TIER2, value=t2.value)

        # Additional Indicators: activity structure
        ai_base = ws_ai.max_row + 1
        ws_ai.cell(row=ai_base, column=1, value=str(ar.title)).fill = _ORANGE
        ws_ai.cell(row=ai_base + 1, column=1, value="Land Uses Targeted (ha)").fill = _BLUE
        ws_ai.cell(row=ai_base + 2, column=1, value="With project").fill = _BEIGE
        ws_ai.cell(row=ai_base + 3, column=1, value="Without project").fill = _BEIGE

        # with_row: the "With project" row – fixed, never moves
        # wo_row: the "Without project" row – shifts down as w rows are inserted
        with_row = ai_base + 2
        wo_row = ai_base + 3

        # Write modules
        for mr in ar.module_results:
            self._write_module_results(ws_res, mr, dur)
            self._write_module_metadata(ws_meta, mr)

            # Additional Indicators: replicate original insert_rows logic.
            # w rows: always inserted right after "With project" (with_row is fixed).
            # Each insert shifts "Without project" and all existing wo data down by 1.
            for label, values in mr.additional_indicator_rows_w:
                ws_ai.insert_rows(with_row + 1)
                ws_ai.cell(with_row + 1, 1, label)
                for i, v in enumerate(values[:dur]):
                    ws_ai.cell(with_row + 1, i + DATA_COL_START, v)
                wo_row += 1  # "Without project" shifted down

            # wo rows: always inserted right after "Without project" (wo_row).
            # wo_row stays fixed – wo data accumulates below it in reverse module order.
            for label, values in mr.additional_indicator_rows_wo:
                ws_ai.insert_rows(wo_row + 1)
                ws_ai.cell(wo_row + 1, 1, label)
                for i, v in enumerate(values[:dur]):
                    ws_ai.cell(wo_row + 1, i + DATA_COL_START, v)
                # wo_row does NOT change: next insert also goes right after "Without project"

        # Results: write total_emissions values on the activity title row
        for i in range(dur):
            v = ar.total_emissions[i] if i < len(ar.total_emissions) else 0
            ws_res.cell(row=activity_row, column=i + DATA_COL_START, value=v).fill = _BEIGE

    def _write_module_results(
        self, ws_res: pxl.worksheet.worksheet.Worksheet, mr: ModuleResult, dur: int
    ) -> None:
        # Module title row
        title_row = ws_res.max_row + 1
        ws_res.cell(row=title_row, column=1, value=str(mr.title)).fill = _BLUE

        # Result data rows
        for rr in mr.result_rows:
            title_row += 1
            ws_res.cell(row=title_row, column=1, value=rr.label)
            vals = rr.yearly_values
            for i in range(dur):
                ws_res.cell(row=title_row, column=i + DATA_COL_START,
                            value=vals[i] if i < len(vals) else 0)

    def _write_module_metadata(
        self, ws_meta: pxl.worksheet.worksheet.Worksheet, mr: ModuleResult
    ) -> None:
        section_base = ws_meta.max_row + 1

        if mr.metadata_section_title:
            ws_meta.cell(
                row=section_base, column=1, value=mr.metadata_section_title
            ).fill = _BLUE

        for mw in mr.metadata_writes:
            row = section_base + mw.row_offset
            cell = ws_meta.cell(row=row, column=mw.col, value=mw.value)
            if mw.fill:
                cell.fill = _FILL_MAP.get(mw.fill, _BEIGE)

    # -------------------------------------------------------------------------
    # Project-level summary (Results rows 2–10)
    # -------------------------------------------------------------------------

    def _write_results_summary(self, wb: pxl.Workbook) -> None:
        ws_res = wb[WS_RESULTS]
        agg = self.result.aggregated

        yearly_balance = list(map(sum, zip_longest(
            agg.biomass_co2, agg.soil_co2, agg.other_co2,
            agg.ch4, agg.n2o, agg.other_ghgs,
            fillvalue=0,
        )))
        cumulative_balance = list(np.cumsum(yearly_balance))

        chf = self.result.cumulative_hectares_yearly
        hhf = self.result.cumulative_heads_yearly
        ctf = self.result.cumulative_catch_yearly
        for i, _year in enumerate(range(self.result.start_year, self.result.last_year)):
            col = i + DATA_COL_START
            ws_res.cell(row=ROW_CUMULATIVE_BALANCE, column=col, value=cumulative_balance[i])
            ws_res.cell(row=ROW_YEARLY_BALANCE, column=col, value=yearly_balance[i])
            ws_res.cell(row=ROW_BIOMASS_CO2, column=col, value=agg.biomass_co2[i])
            ws_res.cell(row=ROW_SOIL_CO2, column=col, value=agg.soil_co2[i])
            ws_res.cell(row=ROW_OTHER_CO2, column=col, value=agg.other_co2[i])
            ws_res.cell(row=ROW_CH4, column=col, value=agg.ch4[i])
            ws_res.cell(row=ROW_N2O, column=col, value=agg.n2o[i])
            ws_res.cell(row=ROW_OTHER_GHGS, column=col, value=agg.other_ghgs[i])
            hv = chf[i] if i < len(chf) else 0
            ws_res.cell(row=ROW_CUMULATIVE_HECTARES, column=col, value=hv).fill = _BEIGE
            hh = hhf[i] if i < len(hhf) else 0
            ws_res.cell(row=ROW_CUMULATIVE_HEADS, column=col, value=hh).fill = _BEIGE
            ct = ctf[i] if i < len(ctf) else 0
            ws_res.cell(row=ROW_CUMULATIVE_CATCH, column=col, value=ct).fill = _BEIGE

    # -------------------------------------------------------------------------
    # Shadow Price of Carbon sheet
    # -------------------------------------------------------------------------

    def _write_shadow_price(self, wb: pxl.Workbook) -> None:
        ws = wb[WS_SHADOW_PRICE]

        # Row 1 header (4 cols merged look via empty cells with same fill)
        for col in range(1, 5):
            ws.cell(
                row=1, column=col,
                value="Economic value of GHG fluxes" if col == 1 else "",
            ).fill = _ORANGE

        ws.cell(row=2, column=1, value="Year")
        ws.cell(row=3, column=1, value="Without (tCO2-eq)")
        ws.cell(row=4, column=1, value="Low Boundary")
        ws.cell(row=5, column=1, value="High Boundary")
        ws.cell(row=7, column=1, value="")
        ws.cell(row=8, column=1, value="With (tCO2-eq)")
        ws.cell(row=9, column=1, value="Low Boundary")
        ws.cell(row=10, column=1, value="High Boundary")

        for i, spr in enumerate(self.result.shadow_price_rows):
            col = i + DATA_COL_START
            ws.cell(row=SPC_ROW_YEAR, column=col, value=spr.year)
            ws.cell(row=SPC_ROW_WO, column=col, value=spr.yearly_balance_wo)
            ws.cell(row=SPC_ROW_WITH, column=col, value=spr.yearly_balance_w)
            if spr.sp_wo_min is not None:
                ws.cell(row=SPC_ROW_WO_LOW, column=col, value=spr.sp_wo_min)
                ws.cell(row=SPC_ROW_WO_HIGH, column=col, value=spr.sp_wo_max)
                ws.cell(row=SPC_ROW_WITH_LOW, column=col, value=spr.sp_w_min)
                ws.cell(row=SPC_ROW_WITH_HIGH, column=col, value=spr.sp_w_max)

        # NPV section
        npv_start = SPC_NPV_START_ROW
        for col in range(1, 5):
            ws.cell(
                row=npv_start, column=col,
                value="SPC with Net Present Value discounting" if col == 1 else "",
            ).fill = _ORANGE

        ws.cell(row=npv_start + 1, column=1, value="Insert rate of discounting")
        rate_cell = ws.cell(row=npv_start + 1, column=3, value=0.05)
        rate_cell.number_format = "0%"
        rate_ref = f"$C${npv_start + 1}"

        npv_year_row = npv_start + 3
        ws.cell(row=npv_year_row, column=1, value="Year")
        ws.cell(row=npv_year_row + 1, column=1, value="Without (tCO2-eq)")
        ws.cell(row=npv_year_row + 2, column=1, value="Low Boundary $")
        ws.cell(row=npv_year_row + 3, column=1, value="High Boundary $")
        ws.cell(row=npv_year_row + 5, column=1, value="With (tCO2-eq)")
        ws.cell(row=npv_year_row + 6, column=1, value="Low Boundary $")
        ws.cell(row=npv_year_row + 7, column=1, value="High Boundary $")

        spc_to_npv = {
            SPC_ROW_WO: npv_year_row + 1,
            SPC_ROW_WO_LOW: npv_year_row + 2,
            SPC_ROW_WO_HIGH: npv_year_row + 3,
            SPC_ROW_WITH: npv_year_row + 5,
            SPC_ROW_WITH_LOW: npv_year_row + 6,
            SPC_ROW_WITH_HIGH: npv_year_row + 7,
        }

        num_years = self.result.last_year - self.result.start_year
        for i in range(num_years):
            col = i + DATA_COL_START
            col_letter = get_column_letter(col)
            ws.cell(row=npv_year_row, column=col, value=f"={col_letter}2")
            for spc_row, npv_row in spc_to_npv.items():
                ws.cell(
                    row=npv_row, column=col,
                    value=f"={col_letter}{spc_row}/(1+{rate_ref})^({col_letter}$2-$B$2)",
                )

        # Nominal SPC table
        ws.cell(
            row=SPC_NOMINAL_ROW, column=1,
            value="Shadow Price of Carbon, nominal (2017 $US) - World Bank*",
        ).fill = _ORANGE
        for col in range(2, 8):
            ws.cell(row=SPC_NOMINAL_ROW, column=col, value="").fill = _ORANGE
        ws.cell(row=SPC_NOMINAL_ROW + 1, column=1, value="Year")
        ws.cell(row=SPC_NOMINAL_ROW + 2, column=1, value="Low 2017")
        ws.cell(row=SPC_NOMINAL_ROW + 3, column=1, value="High")

        nominal = self.result.nominal_shadow_prices
        extra = self.result.extra_shadow_prices
        for i, sp in enumerate(nominal + extra):
            col = i + DATA_COL_START
            ws.cell(row=SPC_NOMINAL_ROW + 1, column=col, value=sp.year)
            ws.cell(row=SPC_NOMINAL_ROW + 2, column=col, value=round(sp.min_value, 2))
            ws.cell(row=SPC_NOMINAL_ROW + 3, column=col, value=round(sp.max_value, 2))
            if i >= len(nominal):
                ws.cell(row=SPC_NOMINAL_ROW + 1, column=col).fill = _RED
                ws.cell(row=SPC_NOMINAL_ROW + 2, column=col).fill = _RED
                ws.cell(row=SPC_NOMINAL_ROW + 3, column=col).fill = _RED

        ws.cell(
            row=SPC_NOMINAL_ROW + 5, column=1,
            value=(
                "* The shadow price of carbon (SPC) beyond 2050 is calculated by applying a 2.25% "
                "annual increase in the SPC values starting from 2050, as per 2024 Guidance Note on "
                "Shadow Price of Carbon in Economic Analysis"
            ),
        )

    # -------------------------------------------------------------------------
    # Inventory sheet
    # -------------------------------------------------------------------------

    def _write_inventory(self, wb: pxl.Workbook) -> None:
        ws_inv = wb[WS_INVENTORY]
        items = self.result.inventory_items

        if not items:
            ws_inv.cell(
                row=2, column=1,
                value="No inventory data available for this project.",
            ).fill = _BEIGE
            return

        for i, item in enumerate(items, start=2):
            ws_inv.cell(row=i, column=INV_COL_ACTIVITY, value=item.activity_name)
            ws_inv.cell(row=i, column=INV_COL_MODULE, value=item.module_name)
            ws_inv.cell(row=i, column=INV_COL_IPCC_CATEGORY, value=item.ipcc_category)
            ws_inv.cell(row=i, column=INV_COL_GAS_TYPE, value=item.gas_type)
            ws_inv.cell(row=i, column=INV_COL_VALUE, value=item.value)
