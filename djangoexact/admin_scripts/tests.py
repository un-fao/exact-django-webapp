import io

from django.test import TestCase, Client, override_settings
from api.models import CustomUser
from minitool.models import ChangeRecord

# The DatabaseConnectionMiddleware calls connections.close_all() after every
# request, which breaks TestCase's transaction-based isolation. We exclude it
# during tests so the test-runner's database connection stays open.
MIDDLEWARE_WITHOUT_DB_CLEANUP = [
    "django.middleware.security.SecurityMiddleware",
    "django.contrib.sessions.middleware.SessionMiddleware",
    "corsheaders.middleware.CorsMiddleware",
    "django.middleware.common.CommonMiddleware",
    "django.middleware.csrf.CsrfViewMiddleware",
    "django.contrib.auth.middleware.AuthenticationMiddleware",
    "django.middleware.locale.LocaleMiddleware",
    "django.contrib.messages.middleware.MessageMiddleware",
    "django.middleware.clickjacking.XFrameOptionsMiddleware",
    "auditlog.middleware.AuditlogMiddleware",
    "simple_history.middleware.HistoryRequestMiddleware",
]


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class AdminScriptsAccessTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.staff_user = CustomUser.objects.create_user(
            email="staff@example.com",
            password="testpass123",
            is_staff=True,
            firebase_uid="staff_uid",
        )
        self.regular_user = CustomUser.objects.create_user(
            email="user@example.com",
            password="testpass123",
            is_staff=False,
            firebase_uid="user_uid",
        )

    def test_dashboard_redirects_unauthenticated(self):
        response = self.client.get("/api/admin-scripts/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin/login/", response.url)

    def test_dashboard_forbidden_for_non_staff(self):
        self.client.login(email="user@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/")
        self.assertEqual(response.status_code, 403)

    def test_dashboard_accessible_for_staff(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Example Script")

    def test_dashboard_shows_compile_scenarios(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/")
        self.assertContains(response, "Compile Scenarios")

    def test_example_script_get(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/example-script/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Example Script")

    def test_example_script_post_with_input(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/example-script/", {"name": "World"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hello, World!")

    def test_example_script_post_empty_input(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/example-script/", {"name": ""})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Please provide a name")


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class ScenarioUtilsTest(TestCase):
    databases = {"default"}

    def setUp(self):
        for i, total in enumerate([-3.0, -2.5, -2.0, -1.5, -1.0]):
            ChangeRecord.objects.create(
                module_type="Grassland",
                region="Central Asia",
                climate="Cool Temperate",
                moisture="Moist",
                soil_type="High Activity Clay",
                total=total,
                field="grassland_management_type",
                from_value="Non-Degraded",
                to_value="Improved Grassland",
                csv_row_data={"row": i},
            )

    def test_stats_for_returns_all_keys(self):
        from admin_scripts.scenario_utils import stats_for

        qs = ChangeRecord.objects.all()
        result = stats_for(qs)
        expected_keys = {
            "count", "sum_total", "mean", "median", "min", "max",
            "std", "q1", "q3", "iqr", "ci_95", "ci_99",
        }
        self.assertEqual(set(result.keys()), expected_keys)

    def test_stats_for_correct_count_and_mean(self):
        from admin_scripts.scenario_utils import stats_for

        qs = ChangeRecord.objects.all()
        result = stats_for(qs)
        self.assertEqual(result["count"], 5)
        self.assertAlmostEqual(result["mean"], -2.0, places=5)
        self.assertAlmostEqual(result["median"], -2.0, places=5)

    def test_stats_for_empty_queryset(self):
        from admin_scripts.scenario_utils import stats_for

        qs = ChangeRecord.objects.none()
        result = stats_for(qs)
        self.assertEqual(result["count"], 0)
        self.assertIsNone(result["mean"])
        self.assertIsNone(result["std"])

    def test_build_scenario_query_basic(self):
        from admin_scripts.scenario_utils import build_scenario_query

        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
        }]
        q = build_scenario_query(changes, {})
        results = ChangeRecord.objects.filter(q)
        self.assertEqual(results.count(), 5)

    def test_build_scenario_query_with_soil_filter(self):
        from admin_scripts.scenario_utils import build_scenario_query

        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
        }]
        global_filters = {"soil_type": ["Sandy"]}
        q = build_scenario_query(changes, global_filters)
        results = ChangeRecord.objects.filter(q)
        self.assertEqual(results.count(), 0)

    def test_build_scenario_query_multiple_changes(self):
        from admin_scripts.scenario_utils import build_scenario_query

        ChangeRecord.objects.create(
            module_type="Annual Cropland",
            region="Central Asia",
            climate="Cool Temperate",
            moisture="Moist",
            soil_type="High Activity Clay",
            total=-0.5,
            field="organic_input_type",
            from_value="Low C input",
            to_value="High C input",
        )
        changes = [
            {
                "module_type": "Grassland",
                "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
                "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            },
            {
                "module_type": "Annual Cropland",
                "start": {"field": "organic_input_type", "value": "Low C input"},
                "end": {"field": "organic_input_type", "value": "High C input"},
            },
        ]
        q = build_scenario_query(changes, {})
        results = ChangeRecord.objects.filter(q)
        self.assertEqual(results.count(), 6)


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class CompileScenariosViewTest(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = Client()
        self.staff_user = CustomUser.objects.create_user(
            email="staff@example.com",
            password="testpass123",
            is_staff=True,
            firebase_uid="staff_uid",
        )
        ChangeRecord.objects.create(
            module_type="Grassland", region="Central Asia", climate="Cool Temperate",
            moisture="Moist", soil_type="High Activity Clay", total=-2.0,
            field="grassland_management_type", from_value="Non-Degraded",
            to_value="Improved Grassland",
        )
        ChangeRecord.objects.create(
            module_type="Grassland", region="Eastern Europe", climate="Warm Temperate",
            moisture="Dry", soil_type="Sandy", total=-1.5,
            field="grassland_management_type", from_value="High Intensity Grazing",
            to_value="Non-Degraded",
        )
        ChangeRecord.objects.create(
            module_type="Annual Cropland", region="Central Asia", climate="Cool Temperate",
            moisture="Moist", soil_type="High Activity Clay", total=-0.5,
            field="organic_input_type", from_value="Low C input", to_value="High C input",
        )

    def test_compile_scenarios_get_returns_form(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Compile Scenarios")
        self.assertContains(response, "Scenario Name")
        self.assertContains(response, "Add Another Change")
        self.assertContains(response, "Run Scenario")

    def test_compile_scenarios_requires_staff(self):
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 302)

    def test_run_scenario_returns_statistics(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")
        self.assertContains(response, "Count")

    def test_run_scenario_no_matching_records(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Nonexistent",
            "scenario-0-change-0-field": "fake_field",
            "scenario-0-change-0-from_value": "A",
            "scenario-0-change-0-to_value": "B",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No matching records")

    def test_run_scenario_with_global_filters(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "global_filter_soil_type": ["Sandy"],
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")

    def test_run_scenario_multiple_changes(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "scenario-0-change-1-module_type": "Annual Cropland",
            "scenario-0-change-1-field": "organic_input_type",
            "scenario-0-change-1-from_value": "Low C input",
            "scenario-0-change-1-to_value": "High C input",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")

    def test_run_scenario_missing_changes(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "at least one change")

    def test_htmx_module_types(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/module-types/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Grassland")
        self.assertContains(response, "Annual Cropland")

    def test_htmx_fields(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/fields/",
            {"change-0-module_type": "Grassland"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "grassland_management_type")

    def test_htmx_fields_requires_module_type(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/fields/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select module type first")

    def test_htmx_values(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/values/",
            {
                "change-0-module_type": "Grassland",
                "change-0-field": "grassland_management_type",
                "index": "0",
                "prefix": "change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Non-Degraded")
        self.assertContains(response, "Improved Grassland")

    def test_htmx_values_requires_both_params(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/values/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select module type and field first")

    def test_htmx_filters(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/filters/",
            {
                "change-0-module_type": "Grassland",
                "index": "0",
                "prefix": "change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Central Asia")
        self.assertContains(response, "Eastern Europe")
        self.assertContains(response, "Cool Temperate")

    def test_htmx_add_change(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-change/",
            {"index": "1"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Change #2")
        self.assertContains(response, "change-1-module_type")

    def test_export_to_excel(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/export/", {
            "scenario-0-scenario_name": "Test Export",
            "scenario-0-category": "Test",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_export_requires_post(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/export/")
        self.assertEqual(response.status_code, 405)

    def test_parse_changes_with_scenario_prefix(self):
        from admin_scripts.views import _parse_changes_from_post
        from django.http import QueryDict

        post = QueryDict(mutable=True)
        post["scenario-0-change-0-module_type"] = "Grassland"
        post["scenario-0-change-0-field"] = "grassland_management_type"
        post["scenario-0-change-0-from_value"] = "Non-Degraded"
        post["scenario-0-change-0-to_value"] = "Improved Grassland"
        post.setlist("scenario-0-change-0-filter-region", ["Central Asia"])

        changes = _parse_changes_from_post(post, prefix="scenario-0-")
        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["module_type"], "Grassland")
        self.assertEqual(changes[0]["start"]["field"], "grassland_management_type")
        self.assertEqual(changes[0]["start"]["value"], "Non-Degraded")
        self.assertEqual(changes[0]["end"]["value"], "Improved Grassland")
        self.assertEqual(changes[0]["filters"]["region"], ["Central Asia"])

    def test_parse_scenarios_from_post(self):
        from admin_scripts.views import _parse_scenarios_from_post
        from django.http import QueryDict

        post = QueryDict(mutable=True)
        post["scenario-0-scenario_name"] = "Scenario A"
        post["scenario-0-category"] = "Cat A"
        post["scenario-0-change-0-module_type"] = "Grassland"
        post["scenario-0-change-0-field"] = "grassland_management_type"
        post["scenario-0-change-0-from_value"] = "Non-Degraded"
        post["scenario-0-change-0-to_value"] = "Improved Grassland"
        post["scenario-1-scenario_name"] = "Scenario B"
        post["scenario-1-category"] = "Cat B"
        post["scenario-1-change-0-module_type"] = "Annual Cropland"
        post["scenario-1-change-0-field"] = "organic_input_type"
        post["scenario-1-change-0-from_value"] = "Low C input"
        post["scenario-1-change-0-to_value"] = "High C input"

        scenarios = _parse_scenarios_from_post(post)
        self.assertEqual(len(scenarios), 2)
        self.assertEqual(scenarios[0]["scenario_name"], "Scenario A")
        self.assertEqual(scenarios[0]["category"], "Cat A")
        self.assertEqual(len(scenarios[0]["changes"]), 1)
        self.assertEqual(scenarios[0]["changes"][0]["module_type"], "Grassland")
        self.assertEqual(scenarios[1]["scenario_name"], "Scenario B")
        self.assertEqual(len(scenarios[1]["changes"]), 1)
        self.assertEqual(scenarios[1]["changes"][0]["module_type"], "Annual Cropland")

    def test_compile_scenarios_access_forbidden_non_staff(self):
        regular_user = CustomUser.objects.create_user(
            email="regular@example.com", password="testpass123",
            is_staff=False, firebase_uid="regular_uid",
        )
        self.client.login(email="regular@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 403)


class ChangeKeyParsingTest(TestCase):
    def test_extract_change_key_info_old_format(self):
        from admin_scripts.views import _extract_change_key_info

        result = _extract_change_key_info(
            {"change-2-module_type": "Grassland"},
            suffix="module_type"
        )
        self.assertEqual(result, ("Grassland", "2", "change-2-"))

    def test_extract_change_key_info_scenario_format(self):
        from admin_scripts.views import _extract_change_key_info

        result = _extract_change_key_info(
            {"scenario-1-change-3-module_type": "Grassland"},
            suffix="module_type"
        )
        self.assertEqual(result, ("Grassland", "3", "scenario-1-change-3-"))

    def test_extract_change_key_info_no_match(self):
        from admin_scripts.views import _extract_change_key_info

        result = _extract_change_key_info(
            {"unrelated_key": "value"},
            suffix="module_type"
        )
        self.assertIsNone(result)


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class HtmxScenarioPrefixTest(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = Client()
        self.staff_user = CustomUser.objects.create_user(
            email="staff@example.com",
            password="testpass123",
            is_staff=True,
            firebase_uid="staff_uid",
        )
        ChangeRecord.objects.create(
            module_type="Grassland", region="Central Asia", climate="Cool Temperate",
            moisture="Moist", soil_type="High Activity Clay", total=-2.0,
            field="grassland_management_type", from_value="Non-Degraded",
            to_value="Improved Grassland",
        )

    def test_htmx_fields_with_scenario_prefix(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/fields/",
            {"scenario-0-change-0-module_type": "Grassland"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "grassland_management_type")
        self.assertContains(response, 'name="scenario-0-change-0-field"')

    def test_htmx_values_with_scenario_prefix(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/values/",
            {
                "scenario-0-change-0-module_type": "Grassland",
                "scenario-0-change-0-field": "grassland_management_type",
                "index": "0",
                "prefix": "scenario-0-change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Non-Degraded")
        self.assertContains(response, 'name="scenario-0-change-0-from_value"')

    def test_htmx_filters_with_scenario_prefix(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/filters/",
            {
                "scenario-0-change-0-module_type": "Grassland",
                "index": "0",
                "prefix": "scenario-0-change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Central Asia")
        self.assertContains(response, 'name="scenario-0-change-0-filter-region"')

    def test_htmx_add_change_with_scenario_index(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-change/",
            {"index": "1", "scenario_index": "2"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Change #2")
        self.assertContains(response, 'name="scenario-2-change-1-module_type"')

    def test_htmx_run_scenario(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {
                "scenario-0-scenario_name": "Test Scenario",
                "scenario-0-category": "Test",
                "scenario-0-change-0-module_type": "Grassland",
                "scenario-0-change-0-field": "grassland_management_type",
                "scenario-0-change-0-from_value": "Non-Degraded",
                "scenario-0-change-0-to_value": "Improved Grassland",
                "global_filter_soil_type": ["High Activity Clay"],
                "scenario_index": "0",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")
        self.assertContains(response, "Count")

    def test_htmx_run_scenario_no_changes(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {
                "scenario-0-scenario_name": "Empty",
                "scenario-0-category": "Test",
                "scenario_index": "0",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "at least one change")

    def test_htmx_run_scenario_requires_post(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/run-scenario/")
        self.assertEqual(response.status_code, 405)

    def test_htmx_add_scenario(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-scenario/",
            {"index": "1"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'hx-swap-oob="beforeend:#scenario-tabs"')
        self.assertContains(response, 'data-scenario-tab="1"')
        self.assertContains(response, 'onclick="switchScenarioTab(1)"')
        self.assertContains(response, 'data-scenario-panel="1"')
        self.assertContains(response, 'name="scenario-1-scenario_name"')
        self.assertContains(response, 'name="scenario-1-change-0-module_type"')

    def test_export_multi_scenario(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/export/", {
            "scenario-0-scenario_name": "Scenario A",
            "scenario-0-category": "Cat A",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "scenario-1-scenario_name": "Scenario B",
            "scenario-1-category": "Cat B",
            "scenario-1-change-0-module_type": "Grassland",
            "scenario-1-change-0-field": "grassland_management_type",
            "scenario-1-change-0-from_value": "Non-Degraded",
            "scenario-1-change-0-to_value": "Improved Grassland",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        import openpyxl
        buf = io.BytesIO(b"".join(response.streaming_content))
        wb = openpyxl.load_workbook(buf)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Scenario A Changes", wb.sheetnames)
        self.assertIn("Scenario B Changes", wb.sheetnames)
        self.assertNotIn("Scenario A", wb.sheetnames)
        self.assertNotIn("Scenario B", wb.sheetnames)
