import io

from django.test import TestCase, Client, override_settings
from api.models import CustomUser
from minitool.models import ChangeRecord

# The DatabaseConnectionMiddleware calls connections.close_all() after every
# request, which breaks TestCase's transaction-based isolation. We exclude it
# during tests so the test-runner's database connection stays open.
MIDDLEWARE_WITHOUT_DB_CLEANUP = [
    "django.middleware.security.SecurityMiddleware",
    "django.contrib.sessions.middleware.SessionMiddleware",
    "corsheaders.middleware.CorsMiddleware",
    "django.middleware.common.CommonMiddleware",
    "django.middleware.csrf.CsrfViewMiddleware",
    "django.contrib.auth.middleware.AuthenticationMiddleware",
    "django.middleware.locale.LocaleMiddleware",
    "django.contrib.messages.middleware.MessageMiddleware",
    "django.middleware.clickjacking.XFrameOptionsMiddleware",
    "auditlog.middleware.AuditlogMiddleware",
    "simple_history.middleware.HistoryRequestMiddleware",
]


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class AdminScriptsAccessTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.staff_user = CustomUser.objects.create_user(
            email="staff@example.com",
            password="testpass123",
            is_staff=True,
            firebase_uid="staff_uid",
        )
        self.regular_user = CustomUser.objects.create_user(
            email="user@example.com",
            password="testpass123",
            is_staff=False,
            firebase_uid="user_uid",
        )

    def test_dashboard_redirects_unauthenticated(self):
        response = self.client.get("/api/admin-scripts/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin/login/", response.url)

    def test_dashboard_forbidden_for_non_staff(self):
        self.client.login(email="user@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/")
        self.assertEqual(response.status_code, 403)

    def test_dashboard_accessible_for_staff(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Example Script")

    def test_dashboard_shows_compile_scenarios(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/")
        self.assertContains(response, "Compile Scenarios")

    def test_example_script_get(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/example-script/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Example Script")

    def test_example_script_post_with_input(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/example-script/", {"name": "World"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hello, World!")

    def test_example_script_post_empty_input(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/example-script/", {"name": ""})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Please provide a name")


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class ScenarioUtilsTest(TestCase):
    databases = {"default"}

    def setUp(self):
        for i, total in enumerate([-3.0, -2.5, -2.0, -1.5, -1.0]):
            ChangeRecord.objects.create(
                module_type="Grassland",
                region="Central Asia",
                climate="Cool Temperate",
                moisture="Moist",
                soil_type="High Activity Clay",
                total=total,
                field="grassland_management_type",
                from_value="Non-Degraded",
                to_value="Improved Grassland",
                csv_row_data={"row": i},
            )

    def test_stats_for_returns_all_keys(self):
        from admin_scripts.scenario_utils import stats_for

        qs = ChangeRecord.objects.all()
        result = stats_for(qs)
        expected_keys = {
            "count", "sum_total", "mean", "median", "min", "max",
            "std", "q1", "q3", "iqr", "ci_95", "ci_99",
        }
        self.assertEqual(set(result.keys()), expected_keys)

    def test_stats_for_correct_count_and_mean(self):
        from admin_scripts.scenario_utils import stats_for

        qs = ChangeRecord.objects.all()
        result = stats_for(qs)
        self.assertEqual(result["count"], 5)
        self.assertAlmostEqual(result["mean"], -2.0, places=5)
        self.assertAlmostEqual(result["median"], -2.0, places=5)

    def test_stats_for_empty_queryset(self):
        from admin_scripts.scenario_utils import stats_for

        qs = ChangeRecord.objects.none()
        result = stats_for(qs)
        self.assertEqual(result["count"], 0)
        self.assertIsNone(result["mean"])
        self.assertIsNone(result["std"])

    def test_build_scenario_query_basic(self):
        from admin_scripts.scenario_utils import build_scenario_query

        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
        }]
        q = build_scenario_query(changes, {})
        results = ChangeRecord.objects.filter(q)
        self.assertEqual(results.count(), 5)

    def test_build_scenario_query_with_soil_filter(self):
        from admin_scripts.scenario_utils import build_scenario_query

        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
        }]
        global_filters = {"soil_type": ["Sandy"]}
        q = build_scenario_query(changes, global_filters)
        results = ChangeRecord.objects.filter(q)
        self.assertEqual(results.count(), 0)

    def test_build_scenario_query_multiple_changes(self):
        from admin_scripts.scenario_utils import build_scenario_query

        ChangeRecord.objects.create(
            module_type="Annual Cropland",
            region="Central Asia",
            climate="Cool Temperate",
            moisture="Moist",
            soil_type="High Activity Clay",
            total=-0.5,
            field="organic_input_type",
            from_value="Low C input",
            to_value="High C input",
        )
        changes = [
            {
                "module_type": "Grassland",
                "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
                "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            },
            {
                "module_type": "Annual Cropland",
                "start": {"field": "organic_input_type", "value": "Low C input"},
                "end": {"field": "organic_input_type", "value": "High C input"},
            },
        ]
        q = build_scenario_query(changes, {})
        results = ChangeRecord.objects.filter(q)
        self.assertEqual(results.count(), 6)

    def test_coerce_unit_returns_float_for_numeric_string(self):
        from admin_scripts.scenario_utils import _coerce_unit
        self.assertEqual(_coerce_unit("2.5"), 2.5)
        self.assertEqual(_coerce_unit("3"), 3.0)
        self.assertEqual(_coerce_unit(2.0), 2.0)
        self.assertEqual(_coerce_unit(0), 0.0)

    def test_coerce_unit_defaults_to_one_on_invalid_input(self):
        from admin_scripts.scenario_utils import _coerce_unit
        self.assertEqual(_coerce_unit(None), 1.0)
        self.assertEqual(_coerce_unit(""), 1.0)
        self.assertEqual(_coerce_unit("   "), 1.0)
        self.assertEqual(_coerce_unit("abc"), 1.0)

    def test_coerce_unit_clamps_negative_to_one(self):
        from admin_scripts.scenario_utils import _coerce_unit
        self.assertEqual(_coerce_unit("-2"), 1.0)
        self.assertEqual(_coerce_unit(-0.5), 1.0)

    def test_coerce_unit_rejects_non_finite_values(self):
        import math
        from admin_scripts.scenario_utils import _coerce_unit
        self.assertEqual(_coerce_unit(float("nan")), 1.0)
        self.assertEqual(_coerce_unit(float("inf")), 1.0)
        self.assertEqual(_coerce_unit(float("-inf")), 1.0)
        self.assertEqual(_coerce_unit("nan"), 1.0)
        self.assertEqual(_coerce_unit("inf"), 1.0)
        self.assertEqual(_coerce_unit(math.nan), 1.0)

    def test_stats_for_scenario_unit_one_matches_legacy_baseline(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            "unit": "1",
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["count"], 5)
        self.assertAlmostEqual(stats["sum_total"], -10.0, places=5)
        self.assertAlmostEqual(stats["mean"], -2.0, places=5)
        self.assertAlmostEqual(stats["median"], -2.0, places=5)
        self.assertAlmostEqual(stats["min"], -3.0, places=5)
        self.assertAlmostEqual(stats["max"], -1.0, places=5)

    def test_stats_for_scenario_missing_unit_defaults_to_one(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            # no "unit" key at all
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["count"], 5)
        self.assertAlmostEqual(stats["sum_total"], -10.0, places=5)
        self.assertAlmostEqual(stats["mean"], -2.0, places=5)

    def test_stats_for_scenario_blank_unit_defaults_to_one(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            "unit": "",
        }]
        stats = stats_for_scenario(changes, {})
        self.assertAlmostEqual(stats["sum_total"], -10.0, places=5)
        self.assertAlmostEqual(stats["mean"], -2.0, places=5)

    def test_stats_for_scenario_unit_scales_distribution(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            "unit": "2",
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["count"], 5)
        self.assertAlmostEqual(stats["sum_total"], -20.0, places=5)
        self.assertAlmostEqual(stats["mean"], -4.0, places=5)
        self.assertAlmostEqual(stats["median"], -4.0, places=5)
        self.assertAlmostEqual(stats["min"], -6.0, places=5)
        self.assertAlmostEqual(stats["max"], -2.0, places=5)
        # std and CI scale linearly with the multiplier
        unit_one_changes = [dict(changes[0], unit="1")]
        baseline = stats_for_scenario(unit_one_changes, {})
        self.assertAlmostEqual(stats["std"], baseline["std"] * 2, places=5)
        self.assertAlmostEqual(stats["ci_95"], baseline["ci_95"] * 2, places=5)
        self.assertAlmostEqual(stats["ci_99"], baseline["ci_99"] * 2, places=5)

    def test_stats_for_scenario_unit_zero_zeros_out_distribution(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            "unit": "0",
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["count"], 5)
        self.assertEqual(stats["sum_total"], 0.0)
        self.assertEqual(stats["mean"], 0.0)
        self.assertEqual(stats["min"], 0.0)
        self.assertEqual(stats["max"], 0.0)
        self.assertEqual(stats["std"], 0.0)

    def test_stats_for_scenario_negative_unit_treated_as_one(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            "unit": "-2",
        }]
        stats = stats_for_scenario(changes, {})
        # negative unit clamps to 1.0 — sum stays negative, not positive
        self.assertAlmostEqual(stats["sum_total"], -10.0, places=5)

    def test_stats_for_scenario_overlapping_changes_count_once_per_change(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        # Both changes match the same 5 Grassland records (same module_type/field/values).
        # With units 1 and 3, the distribution gets 5 + 5 = 10 scaled values:
        # 5 of [-3, -2.5, -2, -1.5, -1] and 5 of [-9, -7.5, -6, -4.5, -3].
        changes = [
            {
                "module_type": "Grassland",
                "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
                "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
                "unit": "1",
            },
            {
                "module_type": "Grassland",
                "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
                "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
                "unit": "3",
            },
        ]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["count"], 10)
        # sum = -10 * 1 + -10 * 3 = -40
        self.assertAlmostEqual(stats["sum_total"], -40.0, places=5)

    def test_stats_for_scenario_no_changes_returns_empty_stats(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        stats = stats_for_scenario([], {})
        self.assertEqual(stats["count"], 0)
        self.assertEqual(stats["sum_total"], 0.0)
        self.assertIsNone(stats["mean"])
        self.assertIsNone(stats["std"])

    def test_stats_for_scenario_skips_changes_without_module_type(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [
            {
                "module_type": "",
                "start": {"field": "x", "value": "a"},
                "end": {"field": "x", "value": "b"},
                "unit": "2",
            },
            {
                "module_type": "Grassland",
                "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
                "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
                "unit": "1",
            },
        ]
        stats = stats_for_scenario(changes, {})
        # Only the second change contributes; first is skipped.
        self.assertEqual(stats["count"], 5)
        self.assertAlmostEqual(stats["sum_total"], -10.0, places=5)

    def test_stats_for_scenario_no_outliers_in_baseline_fixture(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        # The five-record baseline fixture (-3.0, -2.5, -2.0, -1.5, -1.0) has no
        # values past Q3+1.5*IQR or below Q1-1.5*IQR.
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["outliers_low"], 0)
        self.assertEqual(stats["outliers_high"], 0)

    def test_stats_for_scenario_counts_outliers_outside_iqr_fences(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        # Add two extreme records - one well below Q1 - 1.5*IQR, one above.
        # Baseline fixture has fences ~(-5.0, 1.0); -100 sits far below the
        # low fence and 50 sits far above the high fence regardless of how
        # the extreme records shift Q1/Q3.
        for extreme in (-100.0, 50.0):
            ChangeRecord.objects.create(
                module_type="Grassland",
                region="Central Asia",
                climate="Cool Temperate",
                moisture="Moist",
                soil_type="High Activity Clay",
                total=extreme,
                field="grassland_management_type",
                from_value="Non-Degraded",
                to_value="Improved Grassland",
            )
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["count"], 7)
        self.assertEqual(stats["outliers_low"], 1)
        self.assertEqual(stats["outliers_high"], 1)

    def test_stats_for_scenario_outlier_counts_zero_when_iqr_undefined(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        # With fewer than 4 values, IQR is undefined; outlier counts must be 0.
        ChangeRecord.objects.all().delete()
        ChangeRecord.objects.create(
            module_type="Grassland",
            region="Central Asia",
            climate="Cool Temperate",
            moisture="Moist",
            soil_type="High Activity Clay",
            total=1.0,
            field="grassland_management_type",
            from_value="Non-Degraded",
            to_value="Improved Grassland",
        )
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(stats["count"], 1)
        self.assertEqual(stats["outliers_low"], 0)
        self.assertEqual(stats["outliers_high"], 0)

    def test_stats_for_scenario_per_change_single_change(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [{
            "module_type": "Grassland",
            "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
            "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            "unit": "2",
        }]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(len(stats["per_change"]), 1)
        entry = stats["per_change"][0]
        self.assertEqual(entry["module_type"], "Grassland")
        self.assertEqual(entry["field"], "grassland_management_type")
        self.assertEqual(entry["from_value"], "Non-Degraded")
        self.assertEqual(entry["to_value"], "Improved Grassland")
        self.assertEqual(entry["unit"], 2.0)
        self.assertEqual(entry["count"], 5)
        self.assertAlmostEqual(entry["sum"], -20.0, places=5)
        self.assertAlmostEqual(entry["mean"], -4.0, places=5)
        self.assertEqual(
            entry["label"],
            "Grassland: Non-Degraded → Improved Grassland",
        )

    def test_stats_for_scenario_per_change_two_changes_preserves_order(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        ChangeRecord.objects.create(
            module_type="Annual Cropland",
            region="Central Asia",
            climate="Cool Temperate",
            moisture="Moist",
            soil_type="High Activity Clay",
            total=-0.5,
            field="organic_input_type",
            from_value="Low C input",
            to_value="High C input",
        )
        changes = [
            {
                "module_type": "Annual Cropland",
                "start": {"field": "organic_input_type", "value": "Low C input"},
                "end": {"field": "organic_input_type", "value": "High C input"},
                "unit": "1",
            },
            {
                "module_type": "Grassland",
                "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
                "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
                "unit": "1",
            },
        ]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(len(stats["per_change"]), 2)
        self.assertEqual(stats["per_change"][0]["module_type"], "Annual Cropland")
        self.assertEqual(stats["per_change"][0]["count"], 1)
        self.assertAlmostEqual(stats["per_change"][0]["sum"], -0.5, places=5)
        self.assertEqual(stats["per_change"][1]["module_type"], "Grassland")
        self.assertEqual(stats["per_change"][1]["count"], 5)
        self.assertAlmostEqual(stats["per_change"][1]["sum"], -10.0, places=5)

    def test_stats_for_scenario_per_change_skips_change_without_module_type(self):
        from admin_scripts.scenario_utils import stats_for_scenario
        changes = [
            {"module_type": "", "start": {"field": "", "value": ""}, "end": {"field": "", "value": ""}},
            {
                "module_type": "Grassland",
                "start": {"field": "grassland_management_type", "value": "Non-Degraded"},
                "end": {"field": "grassland_management_type", "value": "Improved Grassland"},
            },
        ]
        stats = stats_for_scenario(changes, {})
        self.assertEqual(len(stats["per_change"]), 1)
        self.assertEqual(stats["per_change"][0]["module_type"], "Grassland")


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class CompileScenariosViewTest(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = Client()
        self.staff_user = CustomUser.objects.create_user(
            email="staff@example.com",
            password="testpass123",
            is_staff=True,
            firebase_uid="staff_uid",
        )
        ChangeRecord.objects.create(
            module_type="Grassland", region="Central Asia", climate="Cool Temperate",
            moisture="Moist", soil_type="High Activity Clay", total=-2.0,
            field="grassland_management_type", from_value="Non-Degraded",
            to_value="Improved Grassland",
        )
        ChangeRecord.objects.create(
            module_type="Grassland", region="Eastern Europe", climate="Warm Temperate",
            moisture="Dry", soil_type="Sandy", total=-1.5,
            field="grassland_management_type", from_value="High Intensity Grazing",
            to_value="Non-Degraded",
        )
        ChangeRecord.objects.create(
            module_type="Annual Cropland", region="Central Asia", climate="Cool Temperate",
            moisture="Moist", soil_type="High Activity Clay", total=-0.5,
            field="organic_input_type", from_value="Low C input", to_value="High C input",
        )

    def test_compile_scenarios_get_returns_form(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Compile Scenarios")
        self.assertContains(response, "Scenario Name")
        self.assertContains(response, "Add Another Change")
        self.assertContains(response, "Run Scenario")

    def test_compile_scenarios_global_region_filter_populated(self):
        """The scenario-level (global) region filter must list the distinct
        regions present in ChangeRecord. Regression: the dropdown used to
        render with no <option> elements because the view never passed
        ``regions`` and the template had no loop. Region is the only global
        filter, so each option appears exactly once."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="global_filter_region"')
        self.assertContains(response, '<option value="Central Asia">Central Asia</option>', count=1)
        self.assertContains(response, '<option value="Eastern Europe">Eastern Europe</option>', count=1)

    def test_compile_scenarios_global_filter_excludes_soil_type(self):
        """soil_type used to be a global filter (with hardcoded selected
        options); it now lives at the per-change level only."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="global_filter_soil_type"')

    def test_compile_scenarios_per_change_climate_filter_populated(self):
        """The per-change Climate dropdown must list distinct ChangeRecord
        climates at initial render."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-0-change-0-filter-climate"')
        self.assertContains(response, '<option value="Cool Temperate">Cool Temperate</option>', count=1)
        self.assertContains(response, '<option value="Warm Temperate">Warm Temperate</option>', count=1)

    def test_compile_scenarios_per_change_moisture_filter_populated(self):
        """The per-change Moisture dropdown must list distinct ChangeRecord
        moisture values at initial render."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-0-change-0-filter-moisture"')
        self.assertContains(response, '<option value="Moist">Moist</option>', count=1)
        self.assertContains(response, '<option value="Dry">Dry</option>', count=1)

    def test_compile_scenarios_per_change_soil_type_filter_populated(self):
        """The per-change Soil Type dropdown must list distinct ChangeRecord
        soil_type values at initial render."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-0-change-0-filter-soil_type"')
        self.assertContains(response, '<option value="High Activity Clay">High Activity Clay</option>', count=1)
        self.assertContains(response, '<option value="Sandy">Sandy</option>', count=1)

    def test_htmx_add_change_populates_filter_dropdowns(self):
        """htmx_add_change must pass climate/moisture/soil_type choices so the
        new change panel's filter dropdowns aren't blank when the user clicks
        + Add Another Change. Region is global only and not in the panel."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-change/",
            {"index": "1", "scenario_index": "0"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-0-change-1-filter-climate"')
        self.assertContains(response, 'name="scenario-0-change-1-filter-moisture"')
        self.assertContains(response, 'name="scenario-0-change-1-filter-soil_type"')
        self.assertNotContains(response, 'name="scenario-0-change-1-filter-region"')
        self.assertContains(response, '<option value="Cool Temperate">Cool Temperate</option>')
        self.assertContains(response, '<option value="Moist">Moist</option>')
        self.assertContains(response, '<option value="High Activity Clay">High Activity Clay</option>')

    def test_htmx_add_scenario_populates_filter_dropdowns(self):
        """htmx_add_scenario must pass climate/moisture/soil_type choices so
        the new scenario's change panel filter dropdowns aren't blank when the
        user clicks + Add Scenario."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-scenario/",
            {"index": "1"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-1-change-0-filter-climate"')
        self.assertContains(response, 'name="scenario-1-change-0-filter-moisture"')
        self.assertContains(response, 'name="scenario-1-change-0-filter-soil_type"')
        self.assertNotContains(response, 'name="scenario-1-change-0-filter-region"')
        self.assertContains(response, '<option value="Cool Temperate">Cool Temperate</option>')
        self.assertContains(response, '<option value="Moist">Moist</option>')
        self.assertContains(response, '<option value="High Activity Clay">High Activity Clay</option>')

    def test_compile_scenarios_requires_staff(self):
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 302)

    def test_run_scenario_returns_statistics(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")
        self.assertContains(response, "Count")

    def test_run_scenario_no_matching_records(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Nonexistent",
            "scenario-0-change-0-field": "fake_field",
            "scenario-0-change-0-from_value": "A",
            "scenario-0-change-0-to_value": "B",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "not yet computed")

    def test_run_scenario_with_global_filters(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "global_filter_region": ["Central Asia"],
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")

    def test_run_scenario_multiple_changes(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "scenario-0-change-1-module_type": "Annual Cropland",
            "scenario-0-change-1-field": "organic_input_type",
            "scenario-0-change-1-from_value": "Low C input",
            "scenario-0-change-1-to_value": "High C input",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")

    def test_run_scenario_missing_changes(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/htmx/run-scenario/", {
            "scenario_index": "0",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "at least one change")

    def test_htmx_module_types(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/module-types/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Grassland")
        self.assertContains(response, "Annual Cropland")

    def test_htmx_fields(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/fields/",
            {"change-0-module_type": "Grassland"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "grassland_management_type")

    def test_htmx_fields_requires_module_type(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/fields/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select module type first")

    def test_htmx_values(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/values/",
            {
                "change-0-module_type": "Grassland",
                "change-0-field": "grassland_management_type",
                "index": "0",
                "prefix": "change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "from_value")
        self.assertContains(response, "to_value")

    def test_htmx_values_requires_both_params(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/values/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select module type and field first")

    def test_htmx_filters(self):
        """htmx_filters narrows the per-change filter dropdowns to the chosen
        module_type. Returns climate/moisture/soil_type options scoped to the
        module — region is global, not per-change, and must not appear."""
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/filters/",
            {
                "change-0-module_type": "Grassland",
                "index": "0",
                "prefix": "change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="change-0-filter-climate"')
        self.assertContains(response, 'name="change-0-filter-moisture"')
        self.assertContains(response, 'name="change-0-filter-soil_type"')
        self.assertContains(response, "Cool Temperate")
        self.assertContains(response, "Warm Temperate")
        self.assertContains(response, "Moist")
        self.assertContains(response, "Dry")
        self.assertContains(response, "High Activity Clay")
        self.assertContains(response, "Sandy")
        self.assertNotContains(response, 'name="change-0-filter-region"')

    def test_htmx_add_change(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-change/",
            {"index": "1"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Change #2")
        self.assertContains(response, "change-1-module_type")

    def test_export_to_excel(self):
        from openpyxl import load_workbook

        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/export/", {
            "scenario-0-scenario_name": "Test Export",
            "scenario-0-category": "Test",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        body = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(io.BytesIO(body))

        # Run Info first, Summary second, then one detail sheet per scenario.
        self.assertEqual(wb.sheetnames[0], "Run Info")
        self.assertEqual(wb.sheetnames[1], "Summary")
        self.assertIn("Test Export", wb.sheetnames)

        summary = wb["Summary"]
        header = [cell.value for cell in summary[1]]
        for expected in ("Count", "Mean", "Median", "Std Dev"):
            self.assertIn(expected, header)
        self.assertNotIn("Distribution", header)
        # First data row records the scenario.
        row_values = [cell.value for cell in summary[2]]
        self.assertEqual(row_values[header.index("Scenario Name")], "Test Export")
        self.assertEqual(row_values[header.index("Category")], "Test")

        detail = wb["Test Export"]
        cell_values = {cell.value for row in detail.iter_rows() for cell in row}
        for expected in ("STATISTICS", "GLOBAL FILTERS", "CHANGES", "Matched Records"):
            self.assertIn(expected, cell_values)
        self.assertNotIn("DISTRIBUTION", cell_values)

    def test_export_to_excel_sanitizes_and_uniquifies_sheet_names(self):
        from openpyxl import load_workbook

        self.client.login(email="staff@example.com", password="testpass123")
        # Same name twice + a name carrying chars Excel forbids in sheet titles.
        response = self.client.post("/api/admin-scripts/compile-scenarios/export/", {
            "scenario-0-scenario_name": "Soil/Land*Restoration?",
            "scenario-0-category": "Cat",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "scenario-1-scenario_name": "Soil/Land*Restoration?",
            "scenario-1-category": "Cat",
            "scenario-1-change-0-module_type": "Grassland",
            "scenario-1-change-0-field": "grassland_management_type",
            "scenario-1-change-0-from_value": "Non-Degraded",
            "scenario-1-change-0-to_value": "Improved Grassland",
        })
        self.assertEqual(response.status_code, 200)

        body = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(io.BytesIO(body))

        # No sheet name contains a forbidden char, both scenarios got a sheet,
        # the second got a deduplicated " (2)" suffix.
        for name in wb.sheetnames:
            for forbidden in "[]:*?/\\":
                self.assertNotIn(forbidden, name)
        scenario_sheets = [n for n in wb.sheetnames if n not in ("Run Info", "Summary")]
        self.assertEqual(len(scenario_sheets), 2)
        self.assertTrue(any(s.endswith("(2)") for s in scenario_sheets))

    def test_export_requires_post(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/export/")
        self.assertEqual(response.status_code, 405)

    def test_parse_changes_with_scenario_prefix(self):
        from admin_scripts.views import _parse_changes_from_post
        from django.http import QueryDict

        post = QueryDict(mutable=True)
        post["scenario-0-change-0-module_type"] = "Grassland"
        post["scenario-0-change-0-field"] = "grassland_management_type"
        post["scenario-0-change-0-from_value"] = "Non-Degraded"
        post["scenario-0-change-0-to_value"] = "Improved Grassland"
        post.setlist("scenario-0-change-0-filter-climate", ["Cool Temperate"])
        post.setlist("scenario-0-change-0-filter-moisture", ["Moist"])
        post.setlist("scenario-0-change-0-filter-soil_type", ["High Activity Clay"])

        changes = _parse_changes_from_post(post, prefix="scenario-0-")
        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["module_type"], "Grassland")
        self.assertEqual(changes[0]["start"]["field"], "grassland_management_type")
        self.assertEqual(changes[0]["start"]["value"], "Non-Degraded")
        self.assertEqual(changes[0]["end"]["value"], "Improved Grassland")
        self.assertEqual(changes[0]["filters"]["climate"], ["Cool Temperate"])
        self.assertEqual(changes[0]["filters"]["moisture"], ["Moist"])
        self.assertEqual(changes[0]["filters"]["soil_type"], ["High Activity Clay"])
        # Region is global only — never appears in change.filters now.
        self.assertNotIn("region", changes[0]["filters"])

    def test_parse_scenarios_from_post(self):
        from admin_scripts.views import _parse_scenarios_from_post
        from django.http import QueryDict

        post = QueryDict(mutable=True)
        post["scenario-0-scenario_name"] = "Scenario A"
        post["scenario-0-category"] = "Cat A"
        post["scenario-0-change-0-module_type"] = "Grassland"
        post["scenario-0-change-0-field"] = "grassland_management_type"
        post["scenario-0-change-0-from_value"] = "Non-Degraded"
        post["scenario-0-change-0-to_value"] = "Improved Grassland"
        post["scenario-1-scenario_name"] = "Scenario B"
        post["scenario-1-category"] = "Cat B"
        post["scenario-1-change-0-module_type"] = "Annual Cropland"
        post["scenario-1-change-0-field"] = "organic_input_type"
        post["scenario-1-change-0-from_value"] = "Low C input"
        post["scenario-1-change-0-to_value"] = "High C input"

        scenarios = _parse_scenarios_from_post(post)
        self.assertEqual(len(scenarios), 2)
        self.assertEqual(scenarios[0]["scenario_name"], "Scenario A")
        self.assertEqual(scenarios[0]["category"], "Cat A")
        self.assertEqual(len(scenarios[0]["changes"]), 1)
        self.assertEqual(scenarios[0]["changes"][0]["module_type"], "Grassland")
        self.assertEqual(scenarios[1]["scenario_name"], "Scenario B")
        self.assertEqual(len(scenarios[1]["changes"]), 1)
        self.assertEqual(scenarios[1]["changes"][0]["module_type"], "Annual Cropland")

    def test_parse_changes_extracts_unit_field(self):
        from admin_scripts.views import _parse_changes_from_post
        from django.http import QueryDict

        post = QueryDict(mutable=True)
        post["scenario-0-change-0-module_type"] = "Grassland"
        post["scenario-0-change-0-field"] = "grassland_management_type"
        post["scenario-0-change-0-from_value"] = "Non-Degraded"
        post["scenario-0-change-0-to_value"] = "Improved Grassland"
        post["scenario-0-change-0-unit"] = "2.5"

        changes = _parse_changes_from_post(post, prefix="scenario-0-")
        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["unit"], "2.5")

    def test_parse_changes_unit_defaults_to_empty_string_when_missing(self):
        from admin_scripts.views import _parse_changes_from_post
        from django.http import QueryDict

        post = QueryDict(mutable=True)
        post["scenario-0-change-0-module_type"] = "Grassland"
        post["scenario-0-change-0-field"] = "grassland_management_type"
        post["scenario-0-change-0-from_value"] = "Non-Degraded"
        post["scenario-0-change-0-to_value"] = "Improved Grassland"
        # no unit key

        changes = _parse_changes_from_post(post, prefix="scenario-0-")
        self.assertEqual(changes[0]["unit"], "")

    def test_run_scenario_applies_unit_multiplier(self):
        """Posting unit=2 doubles the sum/mean compared to unit=1."""
        self.client.login(email="staff@example.com", password="testpass123")

        base_post = {
            "scenario_index": "0",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
        }

        # Baseline: unit=1 (one matching Grassland record with total=-2.0 in fixtures)
        response_one = self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {**base_post, "scenario-0-change-0-unit": "1"},
        )
        self.assertEqual(response_one.status_code, 200)
        self.assertContains(response_one, "-2.0")

        # unit=2 should scale to -4.0
        response_two = self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {**base_post, "scenario-0-change-0-unit": "2"},
        )
        self.assertEqual(response_two.status_code, 200)
        self.assertContains(response_two, "-4.0")

    def test_compile_scenarios_form_renders_unit_input(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-0-change-0-unit"')
        self.assertContains(response, 'Units')

    def test_htmx_add_change_includes_unit_input(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-change/",
            {"index": "1", "scenario_index": "0"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-0-change-1-unit"')

    def test_export_includes_units_column_in_changes_sheet(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/export/", {
            "scenario-0-scenario_name": "Test Export",
            "scenario-0-category": "Test",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "scenario-0-change-0-unit": "2.5",
        })
        self.assertEqual(response.status_code, 200)

        import openpyxl
        buf = io.BytesIO(b"".join(response.streaming_content))
        wb = openpyxl.load_workbook(buf)
        ws = wb["Test Export Changes"]
        # Header row
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Units", headers)
        units_col = headers.index("Units")
        # Parser stores the raw POSTed string; xlsx cell preserves that string.
        self.assertEqual(ws[2][units_col].value, "2.5")

    def test_export_summary_reflects_unit_scaling(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/export/", {
            "scenario-0-scenario_name": "Scaled",
            "scenario-0-category": "Test",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "scenario-0-change-0-unit": "2",
        })
        self.assertEqual(response.status_code, 200)

        import openpyxl
        buf = io.BytesIO(b"".join(response.streaming_content))
        wb = openpyxl.load_workbook(buf)
        ws = wb["Summary"]
        headers = [cell.value for cell in ws[1]]
        sum_col = headers.index("Sum Total")
        # Fixture has one matching Grassland row with total=-2.0; unit=2 -> sum=-4.0
        self.assertAlmostEqual(ws[2][sum_col].value, -4.0, places=5)

    def test_compile_scenarios_access_forbidden_non_staff(self):
        regular_user = CustomUser.objects.create_user(
            email="regular@example.com", password="testpass123",
            is_staff=False, firebase_uid="regular_uid",
        )
        self.client.login(email="regular@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/")
        self.assertEqual(response.status_code, 403)


class ChangeKeyParsingTest(TestCase):
    def test_extract_change_key_info_old_format(self):
        from admin_scripts.views import _extract_change_key_info

        result = _extract_change_key_info(
            {"change-2-module_type": "Grassland"},
            suffix="module_type"
        )
        self.assertEqual(result, ("Grassland", "2", "change-2-"))

    def test_extract_change_key_info_scenario_format(self):
        from admin_scripts.views import _extract_change_key_info

        result = _extract_change_key_info(
            {"scenario-1-change-3-module_type": "Grassland"},
            suffix="module_type"
        )
        self.assertEqual(result, ("Grassland", "3", "scenario-1-change-3-"))

    def test_extract_change_key_info_no_match(self):
        from admin_scripts.views import _extract_change_key_info

        result = _extract_change_key_info(
            {"unrelated_key": "value"},
            suffix="module_type"
        )
        self.assertIsNone(result)


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class HtmxScenarioPrefixTest(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = Client()
        self.staff_user = CustomUser.objects.create_user(
            email="staff@example.com",
            password="testpass123",
            is_staff=True,
            firebase_uid="staff_uid",
        )
        ChangeRecord.objects.create(
            module_type="Grassland", region="Central Asia", climate="Cool Temperate",
            moisture="Moist", soil_type="High Activity Clay", total=-2.0,
            field="grassland_management_type", from_value="Non-Degraded",
            to_value="Improved Grassland",
        )

    def test_htmx_fields_with_scenario_prefix(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/fields/",
            {"scenario-0-change-0-module_type": "Grassland"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "grassland_management_type")
        self.assertContains(response, 'name="scenario-0-change-0-field"')

    def test_htmx_values_with_scenario_prefix(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/values/",
            {
                "scenario-0-change-0-module_type": "Grassland",
                "scenario-0-change-0-field": "grassland_management_type",
                "index": "0",
                "prefix": "scenario-0-change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "from_value")
        self.assertContains(response, 'name="scenario-0-change-0-from_value"')

    def test_htmx_filters_with_scenario_prefix(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/filters/",
            {
                "scenario-0-change-0-module_type": "Grassland",
                "index": "0",
                "prefix": "scenario-0-change-0-",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="scenario-0-change-0-filter-climate"')
        self.assertContains(response, 'name="scenario-0-change-0-filter-moisture"')
        self.assertContains(response, 'name="scenario-0-change-0-filter-soil_type"')
        self.assertContains(response, "Cool Temperate")
        self.assertNotContains(response, 'name="scenario-0-change-0-filter-region"')

    def test_htmx_add_change_with_scenario_index(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-change/",
            {"index": "1", "scenario_index": "2"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Change #2")
        self.assertContains(response, 'name="scenario-2-change-1-module_type"')

    def test_htmx_run_scenario(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {
                "scenario-0-scenario_name": "Test Scenario",
                "scenario-0-category": "Test",
                "scenario-0-change-0-module_type": "Grassland",
                "scenario-0-change-0-field": "grassland_management_type",
                "scenario-0-change-0-from_value": "Non-Degraded",
                "scenario-0-change-0-to_value": "Improved Grassland",
                "global_filter_region": ["Central Asia"],
                "scenario_index": "0",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Results")
        self.assertContains(response, "Count")

    def test_htmx_run_scenario_no_changes(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {
                "scenario-0-scenario_name": "Empty",
                "scenario-0-category": "Test",
                "scenario_index": "0",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "at least one change")

    def test_htmx_run_scenario_requires_post(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get("/api/admin-scripts/compile-scenarios/htmx/run-scenario/")
        self.assertEqual(response.status_code, 405)

    def test_htmx_add_scenario(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.get(
            "/api/admin-scripts/compile-scenarios/htmx/add-scenario/",
            {"index": "1"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'hx-swap-oob="beforeend:#scenario-tabs"')
        self.assertContains(response, 'data-scenario-tab="1"')
        self.assertContains(response, 'onclick="switchScenarioTab(1)"')
        self.assertContains(response, 'data-scenario-panel="1"')
        self.assertContains(response, 'name="scenario-1-scenario_name"')
        self.assertContains(response, 'name="scenario-1-change-0-module_type"')

    def test_export_multi_scenario(self):
        self.client.login(email="staff@example.com", password="testpass123")
        response = self.client.post("/api/admin-scripts/compile-scenarios/export/", {
            "scenario-0-scenario_name": "Scenario A",
            "scenario-0-category": "Cat A",
            "scenario-0-change-0-module_type": "Grassland",
            "scenario-0-change-0-field": "grassland_management_type",
            "scenario-0-change-0-from_value": "Non-Degraded",
            "scenario-0-change-0-to_value": "Improved Grassland",
            "scenario-1-scenario_name": "Scenario B",
            "scenario-1-category": "Cat B",
            "scenario-1-change-0-module_type": "Grassland",
            "scenario-1-change-0-field": "grassland_management_type",
            "scenario-1-change-0-from_value": "Non-Degraded",
            "scenario-1-change-0-to_value": "Improved Grassland",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        import openpyxl
        buf = io.BytesIO(b"".join(response.streaming_content))
        wb = openpyxl.load_workbook(buf)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Scenario A Changes", wb.sheetnames)
        self.assertIn("Scenario B Changes", wb.sheetnames)
        self.assertNotIn("Scenario A", wb.sheetnames)
        self.assertNotIn("Scenario B", wb.sheetnames)


@override_settings(MIDDLEWARE=MIDDLEWARE_WITHOUT_DB_CLEANUP)
class HtmxRunScenarioContextTest(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = Client()
        self.staff_user = CustomUser.objects.create_user(
            email="staff@example.com",
            password="testpass123",
            is_staff=True,
            firebase_uid="staff_uid",
        )
        self.client.login(email="staff@example.com", password="testpass123")
        for i, total in enumerate([-3.0, -2.5, -2.0, -1.5, -1.0]):
            ChangeRecord.objects.create(
                module_type="Grassland",
                region="Central Asia",
                climate="Cool Temperate",
                moisture="Moist",
                soil_type="High Activity Clay",
                total=total,
                field="grassland_management_type",
                from_value="Non-Degraded",
                to_value="Improved Grassland",
                csv_row_data={"row": i},
            )

    def _post_run_scenario(self, scenario_index="0", scenario_name="My Scenario"):
        return self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {
                "scenario_index": scenario_index,
                f"scenario-{scenario_index}-scenario_name": scenario_name,
                f"scenario-{scenario_index}-category": "",
                f"scenario-{scenario_index}-change-0-module_type": "Grassland",
                f"scenario-{scenario_index}-change-0-field": "grassland_management_type",
                f"scenario-{scenario_index}-change-0-from_value": "Non-Degraded",
                f"scenario-{scenario_index}-change-0-to_value": "Improved Grassland",
                f"scenario-{scenario_index}-change-0-unit": "1",
            },
        )

    def test_response_contains_data_scenario_result_attribute(self):
        response = self._post_run_scenario(scenario_index="2", scenario_name="Foo")
        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn('data-scenario-result=', body)
        self.assertIn("data-scenario-index='2'", body)

    def test_data_scenario_result_payload_parses_and_has_expected_keys(self):
        import json, html
        response = self._post_run_scenario(scenario_index="0", scenario_name="Foo")
        body = response.content.decode("utf-8")
        marker = "data-scenario-result='"
        start = body.index(marker) + len(marker)
        end = body.index("'", start)
        raw = html.unescape(body[start:end])
        payload = json.loads(raw)
        self.assertEqual(payload["scenario_index"], "0")
        self.assertEqual(payload["scenario_name"], "Foo")
        self.assertIn("statistics", payload)
        self.assertEqual(payload["statistics"]["count"], 5)
        self.assertIn("outliers_low", payload["statistics"])
        self.assertIn("per_change", payload["statistics"])
        self.assertEqual(payload["gaps"], [])
        self.assertIsNone(payload["error"])
        self.assertFalse(payload["not_computed"])

    def test_data_scenario_result_present_when_no_matching_records(self):
        import json, html
        response = self.client.post(
            "/api/admin-scripts/compile-scenarios/htmx/run-scenario/",
            {
                "scenario_index": "0",
                "scenario-0-scenario_name": "Empty",
                "scenario-0-category": "",
                "scenario-0-change-0-module_type": "Grassland",
                "scenario-0-change-0-field": "grassland_management_type",
                "scenario-0-change-0-from_value": "DoesNotExist",
                "scenario-0-change-0-to_value": "AlsoMissing",
                "scenario-0-change-0-unit": "1",
            },
        )
        body = response.content.decode("utf-8")
        self.assertIn("data-scenario-result=", body)
        marker = "data-scenario-result='"
        start = body.index(marker) + len(marker)
        end = body.index("'", start)
        raw = html.unescape(body[start:end])
        payload = json.loads(raw)
        self.assertIn("statistics", payload)
        self.assertEqual(payload["statistics"]["count"], 0)
