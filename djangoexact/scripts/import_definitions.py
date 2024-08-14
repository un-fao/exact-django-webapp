import openpyxl
import os
from django.apps import apps
import logging as log
import api.models as api_models
from re import sub


def snake_case(s):
    # Replace hyphens with spaces, then apply regular expression substitutions for title case conversion
    # and add an underscore between words, finally convert the result to lowercase
    return "_".join(sub("([A-Z][a-z]+)", r" \1", sub("([A-Z]+)", r" \1", s.replace("-", " "))).split()).lower()


# Open the Excel file
wb = openpyxl.load_workbook(os.path.join("scripts", "Ex_Act_Definitions_LM_0108.xlsx"))

# Loop through all sheets in the Excel file
for sheet_name in wb.sheetnames:

    class_name = sheet_name.replace("_", "")

    print(f"Processing sheet {sheet_name}...")

    try:
        model = apps.get_app_config("api").get_model(class_name)
    except LookupError:
        log.warning(f"Model {class_name} not found in the app 'api'")
        continue

    module_type = api_models.ModuleType.objects.get(class_name=class_name)
    try:
        pre_existing_def = api_models.Definition.objects.get(module_type=module_type)
    except api_models.Definition.DoesNotExist:
        pre_existing_def = api_models.Definition(module_type=module_type)

    # Open sheet
    sheet = wb[sheet_name]
    # Get all C and D columns row by row as a list of tuples (if the row is not empty)
    data = [(sheet[f"C{i}"].value, sheet[f"D{i}"].value) for i in range(2, sheet.max_row + 1) if sheet[f"C{i}"].value]

    # Create a dictionary from the data
    definitions = {snake_case(key): value for key, value in data}

    log.debug(f"Definitions: {definitions}")

    # Update the definition object with the new definitions
    pre_existing_def.definitions = definitions
    pre_existing_def.save()
